"""Sheet Enumerator analyzer (spec §4.2).

Collects per-sheet metadata from the loader's **structure mode** workbook
[D3]: sheet name, visibility, used range, max row/column, dimension trust, and
a tabular-candidate guess. Dimensions are taken from structure mode because
read_only dimensions may be reset/unreliable (spec §4.2); when structure-mode
dimensions look untrustworthy this analyzer marks
``used_range_trusted=False`` and records a warning rather than failing (spec
§9). The ``is_tabular`` override [D2] short-circuits the tabular guess.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import TYPE_CHECKING

from openpyxl.utils import get_column_letter

from ..context import InspectionContext
from ..exceptions import InspectorError
from .._time_axis import is_time_axis_value
from ..heuristics import (
    MIN_TABULAR_POPULATED_COLS,
    NON_TABULAR_DENSITY_THRESHOLD,
    NON_TABULAR_SAMPLE_ROWS,
    WIDE_SPARSE_AXIS_LOOKAHEAD_ROWS,
    WIDE_SPARSE_DENSE_ROW_RATIO,
    WIDE_SPARSE_MIN_AXIS_ROWS,
    WIDE_SPARSE_MIN_POPULATED_COLS,
)
from ..models import SheetProfile
from ..options import get_is_tabular_override
from ..pipeline import Analyzer

if TYPE_CHECKING:  # pragma: no cover - typing only
    from openpyxl.worksheet.worksheet import Worksheet

#: Sheets whose used range is at most this many columns wide are treated as
#: non-tabular by the legacy dimension-only fallback (empty sample / sampling
#: error). The content-aware gate counts *populated* columns instead [issue #3].
_MAX_NON_TABULAR_COLS = 1


@dataclass(frozen=True)
class _SampleRow:
    """One row from the bounded non-tabular classification sample."""

    row_number: int
    values: tuple[object, ...]
    filled: int

    @property
    def first_cell(self) -> object | None:
        """Return column A's value from this sampled row."""

        return self.values[0] if self.values else None


@dataclass(frozen=True)
class _DensitySample:
    """Content summary for the top-of-sheet tabular gate sample."""

    populated_cols: int
    populated_rows: int
    filled: int
    rows: tuple[_SampleRow, ...]

    @property
    def density(self) -> float:
        """Return ``filled / (populated_cols * populated_rows)`` safely."""

        if self.populated_cols == 0 or self.populated_rows == 0:
            return 0.0
        return self.filled / (self.populated_cols * self.populated_rows)


def _is_non_empty(value: object) -> bool:
    """True when a sampled cell holds content (not None, not the empty string).

    Matches the Header Locator's notion of an empty cell so the tabular gate and
    header scoring agree on what counts as populated.
    """

    return value is not None and not (isinstance(value, str) and value == "")


class SheetEnumerator(Analyzer):
    """Enumerate sheets and their basic structural metadata (spec §4.2)."""

    def name(self) -> str:
        """Return the analyzer identifier."""

        return "sheet_enumerator"

    def analyze(self, context: InspectionContext) -> InspectionContext:
        """Populate ``workbook_profile.sheets`` from the structure workbook.

        For each worksheet a :class:`SheetProfile` is created with ``name``,
        ``is_visible``, ``used_range``, ``max_row``/``max_col``,
        ``used_range_trusted``, and ``is_tabular_candidate``. The
        ``is_tabular`` override [D2] takes precedence over the heuristic.

        Args:
            context: Shared context carrying a ready :class:`Loader`.

        Returns:
            The same context with ``workbook_profile.sheets`` populated.
        """

        loader = context.loader
        if loader is None:  # pragma: no cover - guarded by pipeline wiring
            context.add_warning(
                "sheet_enumerator: no loader available; skipping enumeration"
            )
            return context

        workbook = loader.structure_workbook()
        profiles: list[SheetProfile] = []
        for worksheet in workbook.worksheets:
            profiles.append(self._profile_sheet(worksheet, context))

        context.workbook_profile.sheets = profiles
        return context

    def _profile_sheet(
        self, worksheet: "Worksheet", context: InspectionContext
    ) -> SheetProfile:
        """Build a :class:`SheetProfile` for a single worksheet."""

        name = worksheet.title
        is_visible = worksheet.sheet_state == "visible"

        max_row, max_col, used_range, trusted = self._dimensions(worksheet)
        if not trusted:
            context.add_warning(
                f"sheet_enumerator: untrusted dimensions for sheet "
                f"{name!r}; used_range marked unreliable"
            )

        is_tabular, provenance = self._is_tabular_candidate(
            context, name, max_row, max_col
        )

        return SheetProfile(
            name=name,
            is_visible=is_visible,
            is_tabular_candidate=is_tabular,
            is_tabular_provenance=provenance,
            used_range=used_range,
            used_range_trusted=trusted,
            max_row=max_row,
            max_col=max_col,
        )

    def _dimensions(
        self, worksheet: "Worksheet"
    ) -> tuple[int, int, str, bool]:
        """Return ``(max_row, max_col, used_range, trusted)`` (1-based).

        Dimensions come from structure mode. They are considered untrusted
        when openpyxl reports ``None`` (which can occur for files with a reset
        dimension record); in that case the values are coerced to ``1`` and the
        used range to the single anchor cell ``A1``.
        """

        max_row = worksheet.max_row
        max_col = worksheet.max_column

        trusted = max_row is not None and max_col is not None
        if max_row is None:
            max_row = 1
        if max_col is None:
            max_col = 1

        used_range = (
            f"A1:{get_column_letter(max_col)}{max_row}"
            if max_row >= 1 and max_col >= 1
            else "A1"
        )
        return max_row, max_col, used_range, trusted

    def _is_tabular_candidate(
        self,
        context: InspectionContext,
        sheet_name: str,
        max_row: int,
        max_col: int,
    ) -> tuple[bool, str]:
        """Decide whether a sheet looks like a data table (spec §4.2) [D4].

        The ``is_tabular`` override [D2] wins outright (``provenance="manual"``).
        Otherwise the top :data:`~excel_inspector.heuristics.NON_TABULAR_SAMPLE_ROWS`
        rows are sampled in data mode and judged on *content*, not on the
        rightmost-column dimension (issue #3 — the legacy ``max_col`` gate was
        sensitive to which column the text started in):

        * an empty sample defers to the legacy dimension rule (data may begin
          below the window; a truly empty sheet stays non-tabular);
        * at most :data:`~excel_inspector.heuristics.MIN_TABULAR_POPULATED_COLS`
          populated columns -> non-tabular (a single-column cover, any offset);
        * >= 2 populated columns but sample density below
          :data:`~excel_inspector.heuristics.NON_TABULAR_DENSITY_THRESHOLD`
          -> non-tabular (a scattered multi-column cover).

        Robustness (spec §6): a loader domain error (:class:`InspectorError` —
        corrupt/encrypted) propagates so the pipeline aborts (consistent with
        ``pipeline.py``); any other sampling failure falls back to the legacy
        dimension rule with a warning so enumeration never breaks.

        Returns:
            ``(is_tabular_candidate, provenance)`` where provenance is
            ``"manual"`` for an override and ``"heuristic"`` otherwise [D2].
        """

        override = get_is_tabular_override(context.options, sheet_name)
        if override is not None:
            return override, "manual"

        try:
            sample = self._sample_density(context, sheet_name)
        except InspectorError:
            # Loader domain errors (corrupt/encrypted) are NOT absorbed: they
            # must abort the pipeline (spec §6/§9, like pipeline.py).
            raise
        except Exception as exc:  # noqa: BLE001 - robustness policy (spec §6)
            context.add_warning(
                f"sheet_enumerator: tabular sampling failed for sheet "
                f"{sheet_name!r} ({exc!r}); falling back to dimension heuristic"
            )
            return self._dims_tabular(max_row, max_col), "heuristic"

        if sample.populated_cols == 0:
            return self._dims_tabular(max_row, max_col), "heuristic"
        if sample.populated_cols <= MIN_TABULAR_POPULATED_COLS:
            return False, "heuristic"
        if sample.density < NON_TABULAR_DENSITY_THRESHOLD:
            if _looks_like_wide_sparse_matrix(sample):
                return True, "heuristic"
            return False, "heuristic"
        return True, "heuristic"

    @staticmethod
    def _dims_tabular(max_row: int, max_col: int) -> bool:
        """Legacy dimension-only tabular rule (pre-issue-#3 fallback).

        A sheet with no usable area, or only a single populated column by
        dimension, is non-tabular; otherwise tabular.
        """

        if max_row < 1 or max_col < 1:
            return False
        return max_col > _MAX_NON_TABULAR_COLS

    def _sample_density(
        self, context: InspectionContext, sheet_name: str
    ) -> _DensitySample:
        """Sample the top rows in data mode and summarize populated content.

        Reads the top :data:`~excel_inspector.heuristics.NON_TABULAR_SAMPLE_ROWS`
        rows of ``sheet_name`` in data mode [D3] and returns a
        :class:`_DensitySample`. ``populated_cols`` is the number of distinct
        columns holding any non-empty cell, ``populated_rows`` the number of
        rows with any non-empty cell, and ``filled`` the total non-empty cell
        count. The bounded row summaries let issue #22's wide sparse matrix
        gate recognize a dense header row plus a date/period axis without
        materializing more than the same top sample window.

        Raises:
            InspectorError: A loader domain error; the caller re-raises it.
            Exception: Any other sampling failure; the caller absorbs it into a
                warning and falls back (spec §6).
        """

        loader = context.loader
        if loader is None:
            raise RuntimeError("no loader available for tabular sampling")
        worksheet = loader.data_workbook()[sheet_name]

        populated_cols: set[int] = set()
        populated_rows = 0
        filled = 0
        sample_rows: list[_SampleRow] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1, max_row=NON_TABULAR_SAMPLE_ROWS, values_only=True
            ),
            start=1,
        ):
            row_has_content = False
            row_filled = 0
            for col_index, value in enumerate(row):
                if _is_non_empty(value):
                    populated_cols.add(col_index)
                    filled += 1
                    row_filled += 1
                    row_has_content = True
            if row_has_content:
                populated_rows += 1
            sample_rows.append(
                _SampleRow(row_number, tuple(row), row_filled)
            )
        return _DensitySample(
            len(populated_cols), populated_rows, filled, tuple(sample_rows)
        )


def _looks_like_wide_sparse_matrix(sample: _DensitySample) -> bool:
    """Return True for wide time-series matrices whose body is sparse.

    Issue #22: BIS ``Quarterly Series`` sheets have a dense header/metadata
    band over hundreds of columns, followed immediately by rows whose first
    column is a date axis while only a handful of series values are populated.
    Whole-sample density is therefore low, but the sheet is clearly tabular.
    """

    if sample.populated_cols < WIDE_SPARSE_MIN_POPULATED_COLS:
        return False

    for index, row in enumerate(sample.rows):
        dense_ratio = row.filled / sample.populated_cols
        if dense_ratio < WIDE_SPARSE_DENSE_ROW_RATIO:
            continue

        below = sample.rows[
            index + 1 : index + 1 + WIDE_SPARSE_AXIS_LOOKAHEAD_ROWS
        ]
        axis_rows = sum(
            1 for below_row in below if is_time_axis_value(below_row.first_cell)
        )
        if axis_rows >= WIDE_SPARSE_MIN_AXIS_ROWS:
            return True
    return False
