"""Plan Aggregator (spec §4.8) [D1][D2][D5].

The aggregator is the **only** place that converts the inspection domain
(openpyxl 1-based, on :class:`SheetProfile`/:class:`ColumnProfile`) into the
loading domain (pandas 0-based, on :class:`ReadPlan`) [D1]. It synthesizes one
:class:`ReadPlan` per tabular sheet, attaches it to the sheet profile, and
applies overrides [D2].

This module is introduced in Phase 1 as a *minimal v1* aggregator: it handles
the simple, common cases and lays down the coordinate-conversion skeleton that
later phases (header locator, boundary detector) feed richer inputs into.

Coordinate-conversion rules [D1] (fully exercised once boundaries/headers are
populated in later phases):

1.  Leading rows above the header (``1 .. header_row-1``) are absorbed into
    ``skiprows`` (0-based).
2.  Subtotal/total/blank rows below the header (``skip_rows``, 1-based) are
    converted to 0-based absolute indices and merged into ``skiprows``.
3.  ``header`` is normalized to its position in the post-skip frame (0 when the
    header sits immediately after all skipped leading rows).
4.  ``nrows`` is the **whole** 1-based inclusive span ``data_end_row -
    data_start_row + 1``; interior subtotal/blank skips are **not** subtracted.
    pandas ``nrows`` counts the original rows consumed *after the header*, and
    interior ``skiprows`` are dropped from the output but still consume the
    nrows budget. Subtracting them shortens the read window so it never reaches
    the last data row (verified against pandas 3.0.3).
5.  Multi-level headers (plan v2 Task 11.2) [D6]: when
    ``is_multi_level_header`` is set and the merged group rows above the leaf
    header form a contiguous band ending at ``header_row - 1``,
    ``ReadPlan.header`` becomes the band's 0-based **absolute** row list (e.g.
    band rows 1-2 -> ``[0, 1]``) and the leading rows above the band are *not*
    absorbed into ``skiprows`` — pandas applies its merged-cell header
    forward-fill at pre-skip absolute rows when ``skiprows`` is a list, so
    absorbing them corrupts the group labels (measured, pandas 3.0.3). pandas
    ignores rows above the first header row, ``nrows`` still counts the rows
    after the *leaf* header row, and positional ``dtype_map`` keys [D5] stay
    position-valid under a list header (all measured). Derivation vetoes keep
    the single leaf header conservatively: a non-contiguous band and a plan
    that needs ``usecols`` (pandas rejects usecols + multi-index header) each
    veto with a warning; a manual header override [D2] vetoes silently (the
    caller's single declared header row is authoritative).

dtype_map [D5] (spec §4.8): the per-column :attr:`ColumnProfile.inferred_type`
values from the Type Profiler are mapped through :data:`_TYPE_TO_DTYPE` to the
``{position_string: pandas_dtype}`` map; ``number`` columns are omitted (pandas
auto-infers), ``"unknown"`` (unprofiled) columns are skipped, and a per-sheet
``dtype_force`` override wins per key. Keys are the 0-based position within the
usecols-selected frame, which equals :attr:`ColumnProfile.index` because
``usecols`` selects exactly the profiled table span.

v1 fallback: when no ``header_row`` is detected and no override is present, the
first row of the used range is assumed to be the header at 0-based position 0,
with ``engine="openpyxl"`` (implementation plan Phase 1).

Headerless override (spec §9, HIGH #3): a :class:`SheetOverride` that explicitly
sets ``header_row=None`` is *not* the same as the detection fallback. The former
declares the sheet has no header at all, so the plan's ``header`` becomes
``None`` (pandas reads no header row; the first data row is preserved) instead
of the fallback ``0``. Such a plan also carries the :data:`_HEADERLESS_NOTE`
advisory (plan v2 Phase 13 Step 2, L6): without a header anchor no column was
profiled, so the skipped dtype inference is made visible instead of silently
yielding an empty ``dtype_map``.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from .analyzers.merge_analyzer import _classify_kind
from .context import InspectionContext
from .models import (
    InspectionOptions,
    MergeRegion,
    ReadPlan,
    SheetProfile,
    TableBlock,
)
from .options import get_dtype_force, get_sheet_override, has_header_override
from .pipeline import Analyzer

#: Mapping from a :attr:`ColumnProfile.inferred_type` to the pandas dtype string
#: recorded in :attr:`ReadPlan.dtype_map` [D5] (spec §4.8). ``number`` is
#: deliberately absent: pandas already infers an int/float dtype, so emitting a
#: key for it only constrains the loader needlessly (spec: "number -> 생략"). The
#: ``"unknown"`` sentinel is likewise absent — an unprofiled column must not be
#: silently typed (it is skipped by :func:`_infer_dtype_map`).
#:
#:   * ``numeric_text`` -> ``"string"``: keep digit strings like ``"007"`` as
#:     text so the leading zeros survive (verified vs pandas 3.0.3).
#:   * ``date``         -> ``"datetime64[ns]"``: a loadable pandas datetime dtype.
#:   * ``text``         -> ``"string"``: the pandas nullable string dtype.
#:   * ``mixed``        -> ``"object"``: heterogeneous values stay as Python
#:     objects (no single coercion is safe).
_TYPE_TO_DTYPE: dict[str, str] = {
    "numeric_text": "string",
    "date": "datetime64[ns]",
    "text": "string",
    "mixed": "object",
}


def to_zero_based(one_based_row: int) -> int:
    """Convert a single 1-based openpyxl row to a 0-based pandas index [D1].

    Args:
        one_based_row: A 1-based row number (must be ``>= 1``).

    Returns:
        The 0-based equivalent.

    Raises:
        ValueError: If ``one_based_row`` is less than 1.
    """

    if one_based_row < 1:
        raise ValueError(f"row must be 1-based (>= 1), got {one_based_row}")
    return one_based_row - 1


def column_range_to_usecols(
    left_col: int | None, right_col: int | None
) -> str | None:
    """Translate a 1-based inclusive column span to a pandas ``usecols`` range.

    Args:
        left_col: Left column boundary (1-based) or ``None``.
        right_col: Right column boundary (1-based) or ``None``.

    Returns:
        An Excel column-letter range string (e.g. ``"C:E"``), or ``None`` when
        either boundary is missing (meaning "all columns").
    """

    if left_col is None or right_col is None:
        return None
    return f"{get_column_letter(left_col)}:{get_column_letter(right_col)}"


def _apply_header_override(
    profile: SheetProfile, options: InspectionOptions | None
) -> None:
    """Apply a manual ``header_row`` override onto ``profile`` in place [D2].

    When the sheet has a :class:`SheetOverride` registered, its ``header_row``
    (possibly ``None``) is authoritative: the aggregator records it on the
    profile with ``header_provenance="manual"`` and ``header_confidence=1.0``,
    so the analyzer's own estimate is bypassed (spec §4.8, §6) [D2].

    Otherwise, when no header was detected upstream (the v1 fallback), the
    profile is annotated with ``header_provenance="default"`` to honestly record
    that the header position is a default assumption rather than a heuristic.
    """

    if has_header_override(options, profile.name):
        override = get_sheet_override(options, profile.name)
        assert override is not None  # guaranteed by has_header_override
        profile.header_row = override.header_row
        profile.header_provenance = "manual"
        profile.header_confidence = 1.0
        profile.needs_manual_header = False
    elif profile.header_provenance != "heuristic":
        # No upstream heuristic estimate -> v1 default-assumption header.
        profile.header_provenance = "default"


def build_read_plan(
    profile: SheetProfile,
    options: InspectionOptions | None = None,
    warnings: list[str] | None = None,
    *,
    band_start_row: int | None = None,
) -> ReadPlan:
    """Build a v1 :class:`ReadPlan` for one sheet (spec §4.8) [D1][D5].

    Performs the single 1-based -> 0-based conversion for the inspector. In
    Phase 1 most boundary fields are unset, so the simple fallback path is
    taken; the conversion skeleton handles the richer inputs that later phases
    populate.

    Args:
        profile: The (partially) populated sheet profile (1-based coords). The
            ``header_row``/``header_provenance``/``header_confidence`` fields may
            be mutated in place when a header override is applied [D2].
        options: Inspection options for override application [D2].
        warnings: Optional accumulator for non-fatal notices (e.g. a stray
            ``skip_row`` discarded because it sits at/above the header).
        band_start_row: First row (1-based) of the enclosing row band, when
            known. A band starts at a non-blank row (Block Segmenter), so a
            band start above a heuristically detected header proves non-empty
            rows were absorbed by Rule 1 — surfaced as a "no silent loss"
            note (spec §8; issue #8). ``None`` (geometry unknown, e.g. a
            direct v1 call) keeps the plan note-free.

    Returns:
        The synthesized :class:`ReadPlan` (0-based coords).
    """

    _apply_header_override(profile, options)

    skiprows: list[int] = []

    # Distinguish two header_row=None situations (spec §9, HIGH #3):
    #   * an *explicit headerless override* (header_row=None was specified) ->
    #     the loaded frame must have no header row (pandas header=None), so the
    #     first data row is NOT consumed as column names.
    #   * a detection *fallback* (no header found, no override) -> v1 assumes
    #     the first row is the header (header=0).
    headerless_override = (
        has_header_override(options, profile.name)
        and profile.header_row is None
    )

    # usecols from the (optional) left/right column boundaries. Computed
    # before the header so the multi-level derivation can veto itself: pandas
    # 3.0.3 raises "cannot specify usecols when specifying a multi-index
    # header" (measured, plan v2 Task 11.2 Step 0).
    usecols = column_range_to_usecols(
        profile.data_left_col, profile.data_right_col
    )

    # Multi-level header derivation (plan v2 Task 11.2 Step 1) [D6]. Only a
    # *heuristic/default* header is expanded: a manual header_row override is
    # authoritative [D2] — the caller declared exactly one header row, so it
    # is never silently widened to a list.
    multi_header_rows: list[int] | None = None
    if (
        profile.is_multi_level_header
        and not headerless_override
        and profile.header_row is not None
        and profile.header_provenance != "manual"
    ):
        multi_header_rows = _multi_level_header_rows(profile, warnings)
        if multi_header_rows is not None and usecols is not None:
            # pandas 3.0.3: usecols + multi-index header raises ValueError
            # (measured). Conservative fallback: single leaf header.
            if warnings is not None:
                warnings.append(
                    f"sheet '{profile.name}': multi-level header detected but "
                    f"pandas cannot combine usecols ({usecols!r}) with a "
                    f"multi-index header; loading with the single leaf header "
                    f"row {profile.header_row}"
                )
            multi_header_rows = None

    # Rule 1: absorb leading rows above the header into skiprows (0-based).
    header: int | list[int] | None
    header_row = profile.header_row
    if headerless_override:
        header = None
    elif multi_header_rows is not None:
        # [D1] conversion: the contiguous 1-based band rows become 0-based
        # **absolute** indices. Rows above the band are NOT absorbed into
        # skiprows: pandas interprets list-header indices post-skip but (with
        # a list skiprows) applies the merged-cell forward fill at *pre-skip*
        # absolute rows, so absorbing the leading rows breaks the group
        # labels into 'Unnamed: N_level_0' gaps (measured, pandas 3.0.3,
        # plan v2 Task 11.2 Step 0). pandas already ignores rows above the
        # first header row, and the [D1] interior-skip rule below discards
        # any skip at/above the leaf header, so absolute == post-skip here.
        header = [to_zero_based(row) for row in multi_header_rows]
    elif header_row is not None and header_row > 1:
        skiprows.extend(range(0, header_row - 1))
        header = 0
    elif header_row is not None:
        header = 0
    else:
        # v1 fallback: assume the first row of the used range is the header.
        header = 0

    # Rule 2: interior subtotal/total/blank rows -> 0-based absolute indices.
    # Only genuine *interior* skips (strictly below the header and within the
    # data region) are honored; a stray skip at/above the header or above the
    # data start would otherwise corrupt the post-skip header normalization
    # (issue #9). The same filtered list feeds the no-silent-loss notes below.
    interior_skips = _interior_skip_rows(profile, warnings)
    for one_based in interior_skips:
        zero_based = to_zero_based(one_based)
        if zero_based not in skiprows:
            skiprows.append(zero_based)
    skiprows.sort()

    # Rule 4: row count over the whole (inclusive) data region.
    nrows = _compute_nrows(profile)

    # dtype_map [D5]: start from the profiled column types, then let any forced
    # dtypes (dtype_force override) win per key (spec §4.8).
    dtype_map = _infer_dtype_map(profile)
    dtype_map.update(get_dtype_force(options, profile.name))

    # notes: body-merge forward-fill recommendations from the Merge Analyzer
    # (spec §4.4) plus as_formula advisories from the Formula Detector
    # (plan v2 Phase 12), then the excluded subtotal/separator rows so a
    # dropped row never leaves the loaded frame silently (spec §8; issue #2).
    # v1 records the recommendations only; acting on them is the loader's job.
    notes = _body_merge_notes(profile) + _formula_notes(profile)
    notes.extend(_excluded_subtotal_notes(profile, interior_skips))
    # Rows the Rule-1 absorption (or pandas' above-header discard, in the
    # multi-level case) dropped above a heuristically detected header must
    # not vanish silently (spec §8; issue #8). A manual header is the
    # caller's explicit choice — no note (issue #2 precedent).
    notes.extend(
        _rows_above_header_notes(profile, band_start_row, multi_header_rows)
    )

    # Headerless visibility (plan v2 Phase 13 Step 2, L6): with an explicit
    # headerless declaration there is no header anchor, so the Boundary
    # Detector and Type Profiler never ran — the columns stay unprofiled and
    # _infer_dtype_map above contributed nothing. Say so on the plan instead
    # of losing the fact silently (a dtype_force override [D5] still applies).
    if headerless_override:
        notes.append(_HEADERLESS_NOTE)

    # Detection-fallback visibility (issue #10): no header was detected and
    # none was declared, yet the Rule-1 fallback above assumed the first row
    # is the header. A needs-manual sheet (memo-like content) and a genuine
    # table with a weak non-string header are indistinguishable at this
    # point, so the load proceeds — but never silently: the assumption is
    # surfaced so consumers can verify or override (spec §8).
    if profile.header_row is None and not headerless_override:
        notes.append(_NEEDS_MANUAL_HEADER_NOTE)

    return ReadPlan(
        sheet_name=profile.name,
        engine="openpyxl",
        header=header,
        usecols=usecols,
        skiprows=skiprows,
        nrows=nrows,
        dtype_map=dtype_map,
        notes=notes,
    )


def _multi_level_header_rows(
    profile: SheetProfile, warnings: list[str] | None
) -> list[int] | None:
    """Derive the contiguous multi-level header band (plan v2 Task 11.2) [D6].

    Returns the **1-based** rows of the full header band — the maximal
    contiguous run of ``kind="header"``-merged rows ending at
    ``header_row - 1``, plus the leaf ``header_row`` itself (the "헤더 위
    병합 행이 헤더 행과 연속" condition, applied as a suffix). The caller
    converts the band to the 0-based ``ReadPlan.header`` list [D1].

    Conservative fallbacks (plan v2 Task 11.2 Step 1; issue #7):

    * No header merge strictly above the leaf header (e.g. a band-scoped
      block whose band merges were all classified ``body``) -> ``None``,
      silently — there is simply no multi-level evidence in this scope.
    * Merged rows above the header exist but none touches ``header_row - 1``
      -> ``None`` plus a warning; the sheet loads with the single leaf
      header.
    * Merged rows disconnected from that contiguous run (typically a merged
      title banner above the band, issue #7) are excluded from the band with
      a warning (spec §8 no silent loss) — NOT the previous all-or-nothing
      veto, which dropped the whole band and broke the vertical-merge anchor
      names into pandas ``Unnamed: N`` labels.

    Args:
        profile: Sheet (or synthetic block) profile with the final
            ``header_row`` and classified ``merges``.
        warnings: Optional accumulator for the non-contiguous / excluded-row
            notices.

    Returns:
        The 1-based band rows (ascending, ending at ``header_row``), or
        ``None`` when no list header should be emitted.
    """

    header_row = profile.header_row
    if header_row is None:  # pragma: no cover - guarded by the caller
        return None

    rows_above: set[int] = set()
    for merge in profile.merges:
        if merge.kind != "header":
            continue
        _, min_row, _, max_row = range_boundaries(merge.range)
        if min_row is None or max_row is None:  # pragma: no cover - defensive
            continue
        if min_row >= header_row:
            # A merge on the header row itself (wide single-level label) adds
            # no extra band.
            continue
        for row in range(min_row, min(max_row, header_row - 1) + 1):
            rows_above.add(row)

    if not rows_above:
        return None
    band_top = header_row
    while band_top - 1 in rows_above:
        band_top -= 1
    if band_top == header_row:
        # No merged row touches the leaf (e.g. merged row 1, bare row 2,
        # leaf row 3): no usable band at all -> single leaf header.
        if warnings is not None:
            warnings.append(
                f"sheet '{profile.name}': multi-level header band above row "
                f"{header_row} is not contiguous (merged header rows "
                f"{sorted(rows_above)}); loading with the single leaf header "
                f"row (conservative)"
            )
        return None
    disconnected = sorted(row for row in rows_above if row < band_top)
    if disconnected and warnings is not None:
        warnings.append(
            f"sheet '{profile.name}': merged header rows {disconnected} are "
            f"separated from the contiguous header band (rows "
            f"{band_top}-{header_row}) and were excluded from the "
            f"multi-level header (likely a title/banner row; issue #7)"
        )
    return list(range(band_top, header_row + 1))


#: Prefix of every body-merge forward-fill note (spec §4.4). Stable so tests and
#: loaders can recognize the recommendation without parsing the whole string.
_BODY_MERGE_NOTE_PREFIX = "body merge "


def _body_merge_notes(profile: SheetProfile) -> list[str]:
    """Build forward-fill recommendation notes from body merges (spec §4.4).

    The Merge Analyzer classifies each merge region as ``header`` or ``body``
    and records them on :attr:`SheetProfile.merges`. A ``body`` merge (a label
    spanning several data rows/cells) must have its top-left value propagated to
    the remaining cells (forward-fill) at load time; v1 only *records* the
    recommendation here as a :attr:`ReadPlan.notes` entry — the actual fill is
    the loader's responsibility (spec §4.4).

    The note format is::

        "body merge <A1-range> -> forward-fill top-left value (spec §4.4)"

    Merge regions are emitted in the order recorded on the profile; openpyxl
    does not guarantee a stable ``merged_cells.ranges`` order, so tests compare
    the resulting notes as a set rather than by sequence.

    Args:
        profile: The sheet profile carrying classified :attr:`SheetProfile.merges`.

    Returns:
        One forward-fill note per ``body`` merge (empty when there are none).
    """

    notes: list[str] = []
    for merge in profile.merges:
        if merge.kind == "body":
            notes.append(
                f"{_BODY_MERGE_NOTE_PREFIX}{merge.range} -> forward-fill "
                f"top-left value (spec §4.4)"
            )
    return notes


#: Prefix of every as_formula advisory note (plan v2 Phase 12). Stable so
#: tests and loaders can recognize the recommendation without parsing the
#: whole string.
_FORMULA_NOTE_PREFIX = "formula column "

#: Note recorded on an explicitly headerless sheet's plan (plan v2 Phase 13
#: Step 2, L6): without a header anchor no boundary/type analysis ran, so the
#: dtype inference was skipped — made visible instead of silently lost. The
#: exact string is a stable contract (the plan text), pinned by tests.
_HEADERLESS_NOTE = "headerless sheet: dtype inference skipped"

_NEEDS_MANUAL_HEADER_NOTE = (
    "header heuristic failed: first row assumed as header; verify or set a "
    "header_row / headerless override"
)


def _formula_notes(profile: SheetProfile) -> list[str]:
    """Build as_formula advisory notes from the detected columns (Phase 12).

    The Formula Detector marks a column ``read_hint="as_formula"`` when its
    formula cells have no cached results — a value-mode (``data_only=True``)
    load then yields only nulls, and the Type Profiler saw the same nulls, so
    its type signal for the column is meaningless. The aggregator therefore
    skips dtype inference for such columns (see :func:`_infer_dtype_map`) and
    records the recommendation here so the plan's consumer knows the column
    must be re-read in formula mode to get anything at all (plan v2 §6
    Step 3).

    The note format is::

        "formula column <index> (<name>): read_hint=as_formula — cached
        results are empty, dtype inference skipped; re-read with openpyxl
        data_only=False to obtain the formula strings (plan v2 Phase 12)"

    Args:
        profile: The sheet (or synthetic block) profile with final columns.

    Returns:
        One advisory note per ``as_formula`` column, ascending column order
        (empty when there are none).
    """

    notes: list[str] = []
    for column in profile.columns:
        if column.read_hint != "as_formula":
            continue
        name_part = f" ({column.name!r})" if column.name else ""
        notes.append(
            f"{_FORMULA_NOTE_PREFIX}{column.index}{name_part}: "
            f"read_hint=as_formula — cached results are empty, dtype "
            f"inference skipped; re-read with openpyxl data_only=False to "
            f"obtain the formula strings (plan v2 Phase 12)"
        )
    return notes


#: Prefix of every excluded-subtotal-row note (issue #2; spec §8 "No silent
#: loss"). Stable so tests and consumers can recognize the advisory without
#: parsing the whole string.
_EXCLUDED_ROW_NOTE_PREFIX = "excluded subtotal/separator row at sheet row "


def _excluded_subtotal_notes(
    profile: SheetProfile, interior_skips: list[int]
) -> list[str]:
    """Build "no silent loss" notes for excluded subtotal/separator rows (issue #2).

    The Boundary Detector drops subtotal/total/low-density rows into
    ``skip_rows`` (recording each one's label on
    :attr:`SheetProfile.subtotal_skip_labels`), and the aggregator converts them
    to ``ReadPlan.skiprows``/``nrows`` so they never reach the loaded frame.
    spec §8 forbids losing them silently, so one note per excluded row is
    recorded here — naming its 1-based sheet row and (when present) its label.

    Only rows that (a) survived the interior-skip filter [D1] (so a row removed
    via ``skip_rows_remove`` [D2] is correctly *not* reported — it is no longer
    excluded) and (b) were heuristic *non-blank* skips (present in
    ``subtotal_skip_labels``) get a note. A manually added skip
    (``skip_rows_add`` [D2]) is the caller's explicit choice and an interior
    blank separator carries no data — neither is in the labels map, so neither
    produces noise.

    The note format is::

        "excluded subtotal/separator row at sheet row <N> (<label>)"

    The ``(<label>)`` suffix is omitted when the excluded row has no leading
    string label (a purely sparse low-density row).

    Args:
        profile: The sheet (or synthetic block) profile carrying
            :attr:`SheetProfile.subtotal_skip_labels`.
        interior_skips: The filtered interior skip rows (1-based, ascending) —
            the same list folded into ``skiprows`` — so the notes match exactly
            what was dropped, in deterministic row order.

    Returns:
        One note per excluded subtotal/separator row (empty when there are none).
    """

    labels = profile.subtotal_skip_labels
    notes: list[str] = []
    for one_based in interior_skips:
        if one_based not in labels:
            continue
        label = labels[one_based]
        suffix = f" ({label})" if label else ""
        notes.append(f"{_EXCLUDED_ROW_NOTE_PREFIX}{one_based}{suffix}")
    return notes


#: Prefix of every dropped-rows-above-header note (issue #8; spec §8 "No
#: silent loss"). Stable so tests and consumers can recognize the advisory
#: without parsing the whole string.
_ROWS_ABOVE_HEADER_NOTE_PREFIX = "rows above detected header not loaded: "


def _rows_above_header_notes(
    profile: SheetProfile,
    band_start_row: int | None,
    multi_header_rows: list[int] | None,
) -> list[str]:
    """Build the "no silent loss" note for rows dropped above the header (issue #8).

    Rule 1 absorbs rows ``1 .. header_row-1`` into ``skiprows`` (and with a
    multi-level header pandas itself discards everything above the first
    header row), so content above the detected header never reaches the
    loaded frame. A band starts at a non-blank row (Block Segmenter), so
    ``band_start_row < first_header_row`` proves at least one non-empty row
    was dropped — spec §8 forbids losing it silently.

    No note is produced when:

    * ``band_start_row`` is ``None`` — band geometry unknown (a direct v1
      :func:`build_read_plan` call); staying silent preserves the v1 plan.
    * ``header_provenance != "heuristic"`` — a manual header_row override
      [D2] is the caller's explicit choice (issue #2 precedent), and the
      detection fallback / headerless paths absorb nothing above a header.
    * the band starts at (or below) the first header row — nothing dropped.

    Args:
        profile: The sheet (or synthetic block) profile with the final
            ``header_row``/``header_provenance``.
        band_start_row: First row (1-based) of the enclosing band, or ``None``.
        multi_header_rows: The promoted multi-level header band rows, or
            ``None`` for a single header. The dropped span ends above
            ``multi_header_rows[0]`` (the band top), not the leaf header.

    Returns:
        A single-note list, or an empty list (the note is per-plan, one span).
    """

    if band_start_row is None or profile.header_provenance != "heuristic":
        return []
    first_header = (
        multi_header_rows[0] if multi_header_rows else profile.header_row
    )
    if first_header is None or band_start_row >= first_header:
        return []

    top, bottom = band_start_row, first_header - 1
    span = f"sheet row {top}" if top == bottom else f"sheet rows {top}-{bottom}"
    return [
        f"{_ROWS_ABOVE_HEADER_NOTE_PREFIX}{span} (header at row "
        f"{first_header}); use a header_row override if these are data rows"
    ]


def _interior_skip_rows(
    profile: SheetProfile, warnings: list[str] | None
) -> list[int]:
    """Return only genuine interior ``skip_rows`` (sorted, 1-based) [D1].

    A ``skip_row`` is *interior* iff it sits strictly below the header row and
    at or after ``data_start_row``. Rows at/above the header or above the data
    start are discarded (with a warning), because folding them into the
    aggregated ``skiprows`` would shift the post-skip frame and silently break
    header normalization (issue #9).

    When boundary fields are unset (e.g. Phase 1), the only constraint applied
    is "strictly below the header"; this keeps the simple v1 path permissive.
    """

    header_row = profile.header_row
    data_start = profile.data_start_row

    kept: list[int] = []
    for one_based in sorted(profile.skip_rows):
        if header_row is not None and one_based <= header_row:
            if warnings is not None:
                warnings.append(
                    f"aggregator: ignoring skip_row {one_based} on sheet "
                    f"{profile.name!r}: it is at/above the header row "
                    f"{header_row} (not an interior skip)"
                )
            continue
        if data_start is not None and one_based < data_start:
            if warnings is not None:
                warnings.append(
                    f"aggregator: ignoring skip_row {one_based} on sheet "
                    f"{profile.name!r}: it is above data_start_row "
                    f"{data_start} (not an interior skip)"
                )
            continue
        kept.append(one_based)
    return kept


def _infer_dtype_map(profile: SheetProfile) -> dict[str, str]:
    """Build the heuristic ``dtype_map`` from the profiled columns [D5].

    The key is the column's **0-based position within the usecols-selected
    frame**, expressed as a string (spec §4.8 / [D5]). Because
    :attr:`ColumnProfile.index` is itself 0-based from the table top-left and
    ``usecols`` selects exactly that table span, the profile index *is* the
    selected-frame position — no offset is needed.

    Only columns with a committed type contribute a key:

    * ``number`` -> no key (pandas infers int/float; constraining it is
      needless).
    * ``"unknown"`` -> no key (an unprofiled column must not be silently typed).
    * ``read_hint="as_formula"`` -> no key (plan v2 Phase 12): the column's
      cached values are empty, so the Type Profiler's signal for it was pure
      nulls — typing it (e.g. as ``string``) would constrain a column whose
      value-mode load is meaningless anyway. The skip is made visible via the
      :func:`_formula_notes` advisory on the plan.
    * everything else maps through :data:`_TYPE_TO_DTYPE`.

    Args:
        profile: The sheet profile carrying :attr:`SheetProfile.columns`.

    Returns:
        A new ``{position_string: pandas_dtype}`` mapping (possibly empty).
    """

    dtype_map: dict[str, str] = {}
    for column in profile.columns:
        if column.read_hint == "as_formula":
            # Phase 12: empty-cache formula column -> dtype inference skipped
            # (the type signal was all-null); advisory recorded in notes.
            continue
        dtype = _TYPE_TO_DTYPE.get(column.inferred_type)
        if dtype is None:
            # number -> omitted; unknown -> skipped; any unmapped type -> skip.
            continue
        dtype_map[str(column.index)] = dtype
    return dtype_map


def build_block_read_plan(
    profile: SheetProfile,
    block: TableBlock,
    options: InspectionOptions | None = None,
    warnings: list[str] | None = None,
    *,
    band_scoped: bool = False,
) -> ReadPlan:
    """Build one block's :class:`ReadPlan` (plan v2 Task 10.2 Step 3) [D1][D5].

    Reuses the v1 coordinate conversion verbatim: a synthetic
    :class:`SheetProfile` is assembled from the block's 1-based boundaries plus
    the sheet's own metadata (name, dimensions, merges) and fed through
    :func:`build_read_plan`. The "absorb every row above the header" rule
    (``1 .. header_row-1`` -> 0-based ``skiprows``) holds regardless of the
    block's vertical position, so a lower block's leading rows — including the
    blank separator run and the upper tables — are skipped wholesale and the
    header normalizes to post-skip ``0`` ([D1] invariant; the pandas round-trip
    is pinned by golden tests, guard 5).

    Band clamp (defense line, W-A review HIGH): no block's plan may read
    beyond its own band. With an unresolved data region (``data_end_row is
    None``) the v1 conversion yields ``nrows=None`` (read to EOF), which from
    a block whose band ends before the sheet does would silently swallow every
    later band into one frame (aggregation duplication). Such a plan is
    clamped to ``nrows = band_end_row - header_row`` (the rows below the
    header *inside* the band). When the band already reaches ``max_row``
    (every single-band corpus sheet), EOF equals the band end, so no clamp is
    applied and the mirror block's plan stays identical to the v1 flat plan.

    Override handling (guard 4): the per-block [D2] semantics — ``header_row``
    and ``skip_rows_add`` applying only to the block containing the absolute
    row — were already folded into the block's fields by the Block Analyzer,
    so the sheet-level header-override application inside
    :func:`build_read_plan` is bypassed (``options=None``). Only the
    position-keyed ``dtype_force`` map [D5] is applied here, where the final
    ``dtype_map`` exists.

    Merge attribution (plan v2 Task 11.1 Step 1): a band-scoped block owns
    only the merge regions whose rows intersect its band, so a body merge in
    one stacked table never leaks a forward-fill note into another block's
    plan. The single-band mirror path keeps every sheet merge, preserving the
    mirror-plan == flat-plan compatibility invariant (the flat
    :func:`build_read_plan` path also sees all sheet merges).

    Merge **re-classification** (adversarial review MEDIUM #1): the Merge
    Analyzer classified every sheet merge against the *sheet* header — i.e.
    ``blocks[0]``'s header after mirroring — so a lower band's group merges
    (e.g. ``상반기``/``하반기`` over its own leaf header) arrive here
    mis-labelled ``kind="body"``. That (a) attaches a bogus forward-fill note
    and (b) silently vetoes the block's multi-level header promotion. The
    band-scoped path therefore re-classifies the attributed merges against
    ``block.header_row`` (same rule as ``merge_analyzer._classify_kind``:
    ``header`` iff the merge's top row is at/above the block header) and
    re-derives the block-local ``is_multi_level_header`` flag from the
    re-classified merges, exactly as ``MergeAnalyzer._classify_sheet`` would
    have with the block's header. The mirror path keeps the sheet
    classification verbatim (compat invariant).

    Args:
        profile: The owning sheet (name / dimensions / merges).
        block: The table block whose plan is built (1-based fields).
        options: Inspection options (``dtype_force`` only) [D2].
        warnings: Optional accumulator for non-fatal notices (stray skips).
        band_scoped: ``True`` when the block came from per-band (multi-band)
            analysis rather than the single-band v1 mirror. Band-scoped 1-column
            blocks get a "verify this is a real table" note (review LOW #7);
            the mirror path stays note-free so the mirror plan equals the v1
            flat plan.

    Returns:
        The block's synthesized :class:`ReadPlan` (0-based coords).
    """

    if band_scoped:
        # Attribution (Task 11.1) + re-classification (review MEDIUM #1):
        # the block owns the merges intersecting its band, classified against
        # its OWN header row — the sheet-level classification anchored on the
        # mirrored blocks[0] header and mislabels a lower band's group merges
        # as 'body', losing the multi-level header silently.
        block_merges, block_multi_level = _classify_block_merges(
            _merges_intersecting_rows(
                profile.merges, block.band_start_row, block.band_end_row
            ),
            block.header_row,
        )
    else:
        # Mirror path: sheet merges/classification verbatim, so the mirror
        # block derives the same plan as the flat path (mirror-plan ==
        # flat-plan invariant; plan v2 Task 11.2).
        block_merges = list(profile.merges)
        block_multi_level = profile.is_multi_level_header
    synthetic = SheetProfile(
        name=profile.name,
        is_visible=profile.is_visible,
        is_tabular_candidate=True,
        used_range=profile.used_range,
        used_range_trusted=profile.used_range_trusted,
        max_row=profile.max_row,
        max_col=profile.max_col,
        header_row=block.header_row,
        header_confidence=block.header_confidence,
        header_provenance=block.header_provenance,
        # Plan v2 Task 11.2 / review MEDIUM #1: the mirror path rides the
        # sheet's multi-level flag (mirror-plan == flat-plan invariant); the
        # band-scoped path re-derives the flag from the block's re-classified
        # merges so a lower band's own group-merge band still promotes to a
        # multi-level header list.
        is_multi_level_header=block_multi_level,
        merges=block_merges,
        data_start_row=block.data_start_row,
        data_end_row=block.data_end_row,
        data_left_col=block.data_left_col,
        data_right_col=block.data_right_col,
        skip_rows=list(block.skip_rows),
        columns=list(block.columns),
        subtotal_skip_labels=dict(block.subtotal_skip_labels),
    )
    plan = build_read_plan(
        synthetic, None, warnings, band_start_row=block.band_start_row
    )
    plan.dtype_map.update(get_dtype_force(options, profile.name))

    # Defense line (W-A review HIGH): an unresolved data region must not read
    # past the band end. nrows counts original rows consumed after the header,
    # so the band cap is band_end_row - header_row [D1]. Skipped when the band
    # already reaches max_row (EOF == band end -> v1 mirror plans unchanged).
    if (
        plan.nrows is None
        and block.data_end_row is None
        and block.header_row is not None
        and (profile.max_row <= 0 or block.band_end_row < profile.max_row)
    ):
        plan.nrows = max(0, block.band_end_row - block.header_row)
        plan.notes.append(
            f"data region unresolved; nrows clamped to the band end "
            f"(rows {block.band_start_row}-{block.band_end_row}) so the "
            f"block never reads beyond its band"
        )

    # Review LOW #7: a band-scoped 1-column block is extractable but is often
    # a multi-row comment/footnote, so flag it for human verification.
    if band_scoped and _block_span_width(profile, block) == 1:
        plan.notes.append("1-column band — verify this is a real table")
    return plan


def _merges_intersecting_rows(
    merges: list[MergeRegion], start_row: int, end_row: int
) -> list[MergeRegion]:
    """Merges whose row range intersects ``start_row .. end_row`` (1-based).

    Plan v2 Task 11.1 Step 1 (block attribution): a merge region belongs to
    the block whose band rows it intersects, so a band-scoped block's plan
    only carries its *own* body-merge forward-fill notes. The A1 range is
    parsed with openpyxl's :func:`range_boundaries` (the merge ranges are
    recorded in A1 notation on :class:`MergeRegion`, spec §5.4).

    Args:
        merges: The sheet's classified merge regions.
        start_row: First band row (1-based, inclusive) [D1].
        end_row: Last band row (1-based, inclusive) [D1].

    Returns:
        The intersecting regions, in their original (deterministic) order.
    """

    kept: list[MergeRegion] = []
    for merge in merges:
        _, min_row, _, max_row = range_boundaries(merge.range)
        if max_row is None or min_row is None:  # pragma: no cover - defensive
            kept.append(merge)
            continue
        if max_row >= start_row and min_row <= end_row:
            kept.append(merge)
    return kept


def _classify_block_merges(
    merges: list[MergeRegion], header_row: int | None
) -> tuple[list[MergeRegion], bool]:
    """Re-classify band-attributed merges against the block's own header.

    Adversarial review MEDIUM #1: ``MergeAnalyzer._classify_sheet`` anchors on
    the *sheet* header row (the mirrored ``blocks[0]`` header), so a lower
    stacked table's group merges arrive ``kind="body"`` even though they sit
    on/above that block's own header. Re-classification applies the exact
    spec §4.4 rule (``merge_analyzer._classify_kind``): ``header`` iff the
    merge's top row is at/above ``header_row``; with an unknown header every
    merge stays the conservative ``body``. The multi-level flag is re-judged
    the same way the Merge Analyzer does — ``True`` iff a header merge spans
    a row strictly above the block header (spec §4.4, §5.2).

    Args:
        merges: The merges attributed to the block (band intersection).
        header_row: The block's resolved header row (1-based), or ``None``.

    Returns:
        ``(reclassified, is_multi_level_header)``. New :class:`MergeRegion`
        instances are returned — the sheet's own classified list is never
        mutated.
    """

    reclassified: list[MergeRegion] = []
    has_band_above_header = False
    for merge in merges:
        _, min_row, _, _ = range_boundaries(merge.range)
        if min_row is None:  # pragma: no cover - defensive
            reclassified.append(MergeRegion(range=merge.range, kind=merge.kind))
            continue
        kind = _classify_kind(min_row, header_row)
        reclassified.append(MergeRegion(range=merge.range, kind=kind))
        if kind == "header" and header_row is not None and min_row < header_row:
            has_band_above_header = True
    return reclassified, has_band_above_header


def _block_span_width(profile: SheetProfile, block: TableBlock) -> int | None:
    """The block's table span width in columns (1-based inclusive) [D1].

    Resolved column boundaries win; a full-width block (boundaries ``None`` by
    the spec §4.5 convention) falls back to the sheet's ``max_col``. ``None``
    when neither is known.
    """

    if block.data_left_col is not None and block.data_right_col is not None:
        return block.data_right_col - block.data_left_col + 1
    if profile.max_col and profile.max_col > 0:
        return profile.max_col
    return None


def _compute_nrows(profile: SheetProfile) -> int | None:
    """Compute ``nrows`` for the whole data region (spec §4.8 rule 4) [D1].

    ``nrows`` is the full 1-based inclusive span
    ``data_end_row - data_start_row + 1``. Interior subtotal/blank ``skip_rows``
    are **not** subtracted: pandas ``nrows`` counts the original rows consumed
    after the header, and interior ``skiprows`` are dropped from the output but
    still consume the budget. Subtracting them shortens the window so it never
    reaches the last data row (verified against pandas 3.0.3).

    Returns ``None`` when the data boundaries are unknown (e.g. Phase 1, or an
    empty / header-only sheet).
    """

    start = profile.data_start_row
    end = profile.data_end_row
    if start is None or end is None:
        return None
    return max(0, end - start + 1)


class PlanAggregator(Analyzer):
    """Aggregate analysis results into per-sheet read plans (spec §4.8)."""

    def name(self) -> str:
        """Return the analyzer identifier."""

        return "plan_aggregator"

    def analyze(self, context: InspectionContext) -> InspectionContext:
        """Attach a :class:`ReadPlan` to every tabular sheet profile and block.

        Non-tabular sheets (``is_tabular_candidate=False``) are excluded from
        loading (spec §9) and receive no read plan.

        All-bands-rejected sheets (issue #10): a multi-band tabular candidate
        whose Block Analyzer pass produced **no** block (every band judged
        non-table) also receives no read plan — the flat fallback would load
        the very rows the per-band warnings declared skipped, contradicting
        them (spec §8, no silent loss). The exclusion is surfaced as a
        warning. Override channel stays authoritative [D2]: a header_row
        override (forced or explicit headerless) or a manual
        ``is_tabular=True`` keeps the v1 flat path.

        Blocks (plan v2 Task 10.2 Step 3): every :class:`TableBlock` gets its
        own plan via :func:`build_block_read_plan`. The sheet's flat plan
        follows the mirror rule — for a multi-band sheet it *is*
        ``blocks[0].read_plan`` (the flat fields already mirror that block);
        for a single-band sheet the flat plan keeps the exact v1 path
        (``build_read_plan`` on the real profile, sheet-level overrides and
        warnings included) and the mirror block's independently-computed plan
        must equal it (compatibility assertion, pinned by tests). Block-plan
        warnings are forwarded only on the multi-band path so the single-band
        mirror never duplicates the flat path's notices.

        Args:
            context: Shared context with enumerated sheets.

        Returns:
            The same context with ``read_plan`` set on tabular sheets/blocks.
        """

        for profile in context.workbook_profile.sheets:
            if not profile.is_tabular_candidate:
                profile.read_plan = None
                continue
            bands = context.row_bands.get(profile.name) or []
            if profile.blocks:
                multi_band = len(bands) >= 2
                for block in profile.blocks:
                    block.read_plan = build_block_read_plan(
                        profile,
                        block,
                        context.options,
                        context.warnings if multi_band else None,
                        band_scoped=multi_band,
                    )
                if multi_band:
                    # Flat mirror: the sheet-level plan is the top-most block's.
                    profile.read_plan = profile.blocks[0].read_plan
                    continue
            elif (
                len(bands) >= 2
                and not has_header_override(context.options, profile.name)
                and profile.is_tabular_provenance != "manual"
            ):
                # issue #10: every band was analyzed and judged non-table, so
                # the flat fallback plan would load rows the Block Analyzer
                # warnings already declared "skipped". Emit no plan; surface
                # the exclusion instead (spec §8). Manual overrides [D2] keep
                # the flat path (guarded above).
                profile.read_plan = None
                context.add_warning(
                    f"plan_aggregator: sheet {profile.name!r}: all "
                    f"{len(bands)} detected bands were judged non-table; no "
                    f"read plan emitted (use a header_row or is_tabular "
                    f"override to force loading)"
                )
                continue
            # Band geometry for the no-silent-loss note (issue #8): the flat
            # plan's enclosing band is the sheet's first (the flat fields
            # mirror blocks[0]); the mirror block passes the same value, so
            # the mirror-plan == flat-plan invariant holds.
            profile.read_plan = build_read_plan(
                profile,
                context.options,
                context.warnings,
                band_start_row=bands[0].start_row if bands else None,
            )
        return context
