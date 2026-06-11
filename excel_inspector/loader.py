"""Mode-aware, read-only workbook loader (spec §4.1) [D3].

The loader abstracts the two openpyxl open modes the inspector needs and owns
the lifetime of every workbook handle:

* **Structure mode** — ``load_workbook(path, read_only=False, data_only=True)``.
  Opened once; yields merge regions (``worksheet.merged_cells``), dimensions,
  and sheet metadata. read_only worksheets lack ``merged_cells``, so merge and
  dimension analysis must use this mode (spec §4.1, §4.4) [D3].
* **Data mode** — ``load_workbook(path, read_only=True, data_only=True)``.
  Streaming, single forward pass over sample rows; cached values only.
* **Formula mode (plan v2 Phase 12)** — ``load_workbook(path, read_only=True,
  data_only=False)``: a separate streaming workbook instance whose cell values
  are the *formula strings* (``"=B2*C2"``) instead of cached results [D6].
  Opened **lazily** — only when the Formula Detector actually runs against a
  workbook that contains formula markup — so formula-free inspections never
  pay the second open (plan v2 §6 Step 1/Step 4).

Lifetime / idempotency: handles are cached per mode and lazily opened. All
handles are released by :meth:`Loader.close` (also driven by the context
manager protocol), so the loader never leaks file handles (spec §4.1, §8) [D3].
The loader never writes to the file, guaranteeing byte/idempotency invariants.

Corruption / encryption: openpyxl raises low-level ``BadZipFile`` /
``InvalidFileException`` for both damaged and password-protected files. The
loader disambiguates *before* opening: a genuine OLE2/CFB container (leading
magic ``d0cf11e0``, also reported by :func:`olefile.isOleFile`) is treated as
encrypted; anything else that fails to open is treated as corrupt. Failures are
translated to :class:`EncryptedWorkbookError` / :class:`CorruptWorkbookError`
(spec §4.1, §9).
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import TYPE_CHECKING, NoReturn

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from .exceptions import CorruptWorkbookError, EncryptedWorkbookError

if TYPE_CHECKING:  # pragma: no cover - typing only
    from types import TracebackType

    from openpyxl.workbook.workbook import Workbook

#: Leading magic bytes of an OLE2 / CFB container (encrypted Office documents).
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _is_encrypted_container(path: Path) -> bool:
    """Whether ``path`` looks like an OLE2-wrapped (encrypted) workbook.

    A genuine password-protected ``.xlsx`` is stored inside an OLE2/CFB
    container whose leading bytes are ``d0cf11e0`` (spec §4.1). A plain or
    truncated zip starts with ``PK\\x03\\x04`` instead. We sniff the header
    rather than relying on optional dependencies so the decision is robust.

    Args:
        path: Path to the candidate workbook file.

    Returns:
        ``True`` if the file begins with the OLE2 magic bytes.
    """

    try:
        with path.open("rb") as handle:
            head = handle.read(len(_OLE2_MAGIC))
    except OSError:  # pragma: no cover - defensive; surfaced as corrupt later
        return False
    return head == _OLE2_MAGIC


class Loader:
    """Owns mode-specific workbook handles for one file (spec §4.1) [D3].

    Use as a context manager so handles are always released::

        with Loader(path) as loader:
            wb = loader.structure_workbook()
            ...

    Or manage the lifetime explicitly and call :meth:`close`.

    Args:
        path: Path (or path-like) to the ``.xlsx`` file to inspect.

    Attributes:
        path: The resolved :class:`pathlib.Path` of the target file.
    """

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        self._structure_wb: "Workbook | None" = None
        self._data_wb: "Workbook | None" = None
        self._formula_wb: "Workbook | None" = None
        self._closed = False

    # -- context manager ---------------------------------------------------

    def __enter__(self) -> "Loader":
        return self

    def __exit__(
        self,
        exc_type: "type[BaseException] | None",
        exc: "BaseException | None",
        tb: "TracebackType | None",
    ) -> None:
        self.close()

    # -- open-error translation -------------------------------------------

    def _raise_open_error(self, exc: Exception) -> NoReturn:
        """Translate a low-level open failure to a domain exception.

        Args:
            exc: The originating openpyxl/zipfile exception.

        Raises:
            EncryptedWorkbookError: When the file is an OLE2 container (a
                genuine password-protected workbook).
            CorruptWorkbookError: For any other unreadable file.
        """

        if _is_encrypted_container(self.path):
            raise EncryptedWorkbookError(
                f"workbook is an OLE2/CFB container, not an OOXML zip: "
                f"{self.path}. This is most likely a password/encryption-"
                f"protected .xlsx, but could also be an unsupported legacy "
                f"binary .xls — the two share the OLE2 'd0cf11e0' magic, so "
                f"this classification carries that uncertainty (spec §4.1)."
            ) from exc
        raise CorruptWorkbookError(
            f"workbook could not be opened (corrupt or not a valid .xlsx): "
            f"{self.path}"
        ) from exc

    def _check_not_closed(self) -> None:
        if self._closed:
            raise RuntimeError("Loader has been closed")

    # -- structure mode ----------------------------------------------------

    def structure_workbook(self) -> "Workbook":
        """Open (once) and return the structure-mode workbook [D3].

        ``read_only=False, data_only=True``: exposes ``merged_cells`` and
        trustworthy dimensions. The handle is cached for the loader's lifetime.

        Returns:
            The structure-mode :class:`openpyxl.workbook.workbook.Workbook`.

        Raises:
            EncryptedWorkbookError: The file is encrypted (spec §4.1).
            CorruptWorkbookError: The file is corrupt / not a valid ``.xlsx``.
        """

        self._check_not_closed()
        if self._structure_wb is None:
            try:
                self._structure_wb = openpyxl.load_workbook(
                    self.path, read_only=False, data_only=True
                )
            except (zipfile.BadZipFile, InvalidFileException, OSError) as exc:
                self._raise_open_error(exc)
        return self._structure_wb

    # -- data mode ---------------------------------------------------------

    def data_workbook(self) -> "Workbook":
        """Open (once) and return the data-mode streaming workbook [D3].

        ``read_only=True, data_only=True``: a forward, single-pass streaming
        reader over cached values for sampling rows. The handle is cached for
        the loader's lifetime.

        Returns:
            The data-mode :class:`openpyxl.workbook.workbook.Workbook`.

        Raises:
            EncryptedWorkbookError: The file is encrypted (spec §4.1).
            CorruptWorkbookError: The file is corrupt / not a valid ``.xlsx``.
        """

        self._check_not_closed()
        if self._data_wb is None:
            try:
                self._data_wb = openpyxl.load_workbook(
                    self.path, read_only=True, data_only=True
                )
            except (zipfile.BadZipFile, InvalidFileException, OSError) as exc:
                self._raise_open_error(exc)
        return self._data_wb

    # -- formula mode (plan v2 Phase 12) ------------------------------------

    def formula_workbook(self) -> "Workbook":
        """Open (once) and return the formula-mode streaming workbook [D6].

        ``read_only=True, data_only=False``: a forward streaming reader whose
        cell values are the stored *formula strings* (``"=B2*C2"``) rather
        than cached results — the instance the Formula Detector samples
        (plan v2 §6, spec §4.7). Like the other modes the handle is cached for
        the loader's lifetime, but it is **lazy by contract**: nothing in the
        pipeline calls this until formula markup has actually been found, so a
        formula-free workbook never opens a third handle (plan v2 §6 Step 4).

        Returns:
            The formula-mode :class:`openpyxl.workbook.workbook.Workbook`.

        Raises:
            EncryptedWorkbookError: The file is encrypted (spec §4.1).
            CorruptWorkbookError: The file is corrupt / not a valid ``.xlsx``.
        """

        self._check_not_closed()
        if self._formula_wb is None:
            try:
                self._formula_wb = openpyxl.load_workbook(
                    self.path, read_only=True, data_only=False
                )
            except (zipfile.BadZipFile, InvalidFileException, OSError) as exc:
                self._raise_open_error(exc)
        return self._formula_wb

    # -- lifetime ----------------------------------------------------------

    def close(self) -> None:
        """Close every open workbook handle (idempotent) [D3].

        Each handle is closed in its own ``try`` so a failure closing one does
        not leak the other: every handle is dropped, ``self._closed`` is set
        unconditionally, and any collected close errors are re-raised together
        as an :class:`ExceptionGroup` only after cleanup is complete. Safe to
        call multiple times; subsequent open requests raise
        :class:`RuntimeError`.
        """

        errors: list[Exception] = []
        for attr in ("_structure_wb", "_data_wb", "_formula_wb"):
            wb = getattr(self, attr)
            if wb is None:
                continue
            try:
                wb.close()
            except Exception as exc:  # noqa: BLE001 - isolate per-handle failure
                errors.append(exc)
            finally:
                setattr(self, attr, None)

        self._closed = True
        if errors:
            raise ExceptionGroup("Loader.close encountered errors", errors)
