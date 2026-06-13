"""Corpus-generation tests (implementation plan §5.1, Phase 0 exit criteria).

These tests pin the fixture-generation contract that later phases depend on:

* Every fixture file is generated.
* Each *openable* sample opens cleanly with openpyxl (structure mode) and
  exposes the 1-based coordinates documented in ``generate.FIXTURES``.
* The *negative* samples (corrupt, encrypted) cannot be opened by openpyxl and
  are distinguishable by their leading magic bytes — the signal the Phase 1
  loader will use to choose ``CorruptWorkbookError`` vs ``EncryptedWorkbookError``
  (loader-level translation is tested in the next phase).
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

#: Optional test-only dependency; skip the two negative-fixture tests that need
#: it rather than failing collection if it is not installed (issue #13).
olefile = pytest.importorskip("olefile")

# Imported via conftest's path-based loader; re-import the same module here so
# the FIXTURES metadata and password constant are available to assertions.
from conftest import _load_generate_module  # type: ignore[import-not-found]

_generate = _load_generate_module()
FIXTURES = _generate.FIXTURES
ENCRYPTED_PASSWORD = _generate.ENCRYPTED_PASSWORD

_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_ZIP_LOCAL_MAGIC = b"PK\x03\x04"


def _zip_members(data: bytes) -> dict[str, bytes]:
    """Return a mapping of archive member name -> payload bytes."""

    import io

    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


@pytest.mark.parametrize(
    "builder_name",
    [
        "build_header_simple",
        "build_offset_plus_subtotals",
        "build_merged_header",
        "build_types_mixed",
        "build_hidden_sheet",
        "build_blank_run_terminates",
    ],
)
def test_builders_are_byte_deterministic(builder_name: str) -> None:
    """Calling a builder twice yields byte-identical .xlsx output (issue #5).

    openpyxl otherwise stamps wall-clock timestamps into document properties
    and per-member zip headers; the generator pins both, so the corpus is
    genuinely deterministic (the docstrings' 'deterministic' claim is true).
    """

    builder = getattr(_generate, builder_name)
    assert builder() == builder()


def test_core_xml_modified_is_pinned() -> None:
    """``docProps/core.xml`` carries the fixed timestamp, not wall-clock (HIGH #1).

    openpyxl overwrites ``dcterms:modified`` with the wall-clock save time even
    when ``wb.properties.modified`` is pinned; the generator rewrites the XML
    payload so the fixed W3CDTF value is present and no current-year timestamp
    leaks in.
    """

    members = _zip_members(_generate.build_header_simple())
    core = members["docProps/core.xml"].decode("utf-8")
    fixed = _generate._FIXED_PROPERTY_TS_W3C  # type: ignore[attr-defined]
    assert f"<dcterms:modified" in core
    assert fixed in core
    # The wall-clock save year must not have leaked into the payload.
    import datetime as _dt

    assert str(_dt.date.today().year) not in core


def test_corpus_is_byte_deterministic_across_two_builds() -> None:
    """Building every fixture twice yields byte-identical members (HIGH #1).

    This is the corpus-wide determinism guarantee: for each openable builder,
    two independent builds must produce identical archive members — including
    ``docProps/core.xml`` — so golden/coordinate tests never see flaky bytes.
    """

    for fixture_id, builder in _generate.BUILDERS.items():
        if not FIXTURES[fixture_id].openable:
            continue
        first = _zip_members(builder())
        second = _zip_members(builder())
        assert first.keys() == second.keys(), fixture_id
        for name in first:
            assert first[name] == second[name], f"{fixture_id}:{name}"


def test_builder_is_flaky_free_over_many_iterations() -> None:
    """High-repetition determinism guard (HIGH #1): 50 builds are identical.

    A tight loop is a strong non-flakiness signal; combined with the
    timing-independent guarantees below it pins the corpus as deterministic.
    """

    reference = _generate.build_header_simple()
    for _ in range(50):
        assert _generate.build_header_simple() == reference


def test_pin_core_props_is_wall_clock_independent() -> None:
    """``_pin_core_props`` collapses any modified/created time to the fixed one.

    This is the *timing-independent* proof of the HIGH #1 fix: two core.xml
    payloads that differ only in their ``dcterms:modified`` / ``dcterms:created``
    wall-clock text must normalize to byte-identical output. It does not rely on
    a test happening to straddle a wall-clock second boundary, so it cannot be
    flaky itself.
    """

    template = (
        b'<cp:coreProperties xmlns:dcterms="http://purl.org/dc/terms/">'
        b'<dcterms:created xsi:type="dcterms:W3CDTF">%s</dcterms:created>'
        b'<dcterms:modified xsi:type="dcterms:W3CDTF">%s</dcterms:modified>'
        b"</cp:coreProperties>"
    )
    a = template % (b"2026-06-10T05:01:04Z", b"2026-06-10T05:01:04Z")
    b = template % (b"2099-12-31T23:59:59Z", b"2030-01-02T03:04:05Z")

    pinned_a = _generate._pin_core_props(a)  # type: ignore[attr-defined]
    pinned_b = _generate._pin_core_props(b)  # type: ignore[attr-defined]
    fixed = _generate._FIXED_PROPERTY_TS_W3C  # type: ignore[attr-defined]

    assert pinned_a == pinned_b
    assert fixed.encode("ascii") in pinned_a
    # The original wall-clock timestamps are gone.
    assert b"2026-06-10T05:01:04Z" not in pinned_a
    assert b"2099-12-31T23:59:59Z" not in pinned_b


def test_all_fixtures_are_generated(fixture_corpus: dict[str, Path]) -> None:
    """Every declared fixture id produces an existing, non-empty file."""

    assert set(fixture_corpus) == set(FIXTURES)
    for fixture_id, path in fixture_corpus.items():
        assert path.exists(), f"{fixture_id} not generated at {path}"
        assert path.stat().st_size > 0, f"{fixture_id} is empty"
        assert path.name == FIXTURES[fixture_id].filename


def test_openable_fixtures_open_with_openpyxl(openable_fixture: Path) -> None:
    """Each openable fixture opens in structure mode and has at least one sheet.

    Uses the same open mode the loader's structure mode uses [D3]
    (``read_only=False, data_only=True``); handle is explicitly closed.
    """

    wb = openpyxl.load_workbook(
        openable_fixture, read_only=False, data_only=True
    )
    try:
        assert wb.sheetnames, f"{openable_fixture.name} has no sheets"
    finally:
        wb.close()


def test_header_simple_structure(fixture_path) -> None:
    """header_simple: header row 1, 5 data rows, columns A-D, no merges."""

    wb = openpyxl.load_workbook(fixture_path("header_simple"), data_only=True)
    try:
        ws = wb["Sheet1"]
        assert ws.max_row == 6  # header + 5 data rows (1-based)
        assert ws.max_column == 4
        assert not list(ws.merged_cells.ranges)
        assert ws["A1"].value == "name"
        assert ws["A2"].value == "Alice"
    finally:
        wb.close()


def test_header_offset_structure(fixture_path) -> None:
    """header_offset: title rows 1-3 then header at row 4."""

    wb = openpyxl.load_workbook(fixture_path("header_offset"), data_only=True)
    try:
        ws = wb["Sheet1"]
        assert ws["A1"].value == "월간 판매 보고서"
        assert ws["A4"].value == "product"  # header is on 1-based row 4
        assert ws["A5"].value == "Widget"  # first data row
        assert ws.max_row == 9
        assert ws.max_column == 4
    finally:
        wb.close()


def test_offset_plus_subtotals_coordinates(fixture_path) -> None:
    """offset_plus_subtotals [D1]: header row 4, subtotals 8/12, total 13."""

    wb = openpyxl.load_workbook(
        fixture_path("offset_plus_subtotals"), data_only=True
    )
    try:
        ws = wb["Sheet1"]
        assert ws["A4"].value == "dept"  # header (1-based row 4)
        assert ws["A5"].value == "영업"  # first data row
        assert ws["A8"].value == "소계"  # subtotal (skip_rows)
        assert ws["A12"].value == "소계"  # subtotal (skip_rows)
        assert ws["A13"].value == "합계"  # grand total (skip_rows)
        assert ws.max_row == 13
        assert ws.max_column == 4
    finally:
        wb.close()


def test_merged_header_has_header_and_body_merge(fixture_path) -> None:
    """merged_header: header merge A1:B1 and body merge A6:A7 both present."""

    wb = openpyxl.load_workbook(fixture_path("merged_header"), data_only=True)
    try:
        ws = wb["Sheet1"]
        ranges = {str(m) for m in ws.merged_cells.ranges}
        assert ranges == {"A1:B1", "A6:A7"}
        assert ws["A1"].value == "이름"
        assert ws["C1"].value == "점수"
    finally:
        wb.close()


def test_multi_level_header_has_two_group_merges(fixture_path) -> None:
    """multi_level_header: two row-1 group merges over a row-2 leaf header."""

    wb = openpyxl.load_workbook(
        fixture_path("multi_level_header"), data_only=True
    )
    try:
        ws = wb["Sheet1"]
        ranges = {str(m) for m in ws.merged_cells.ranges}
        assert ranges == {"A1:B1", "C1:D1"}
        assert ws["A1"].value == "상반기"
        assert ws["A2"].value == "1월"  # leaf header on 1-based row 2
    finally:
        wb.close()


def test_types_mixed_storage_kinds(fixture_path) -> None:
    """types_mixed: numeric_text stored as str, date cells, mixed column."""

    wb = openpyxl.load_workbook(fixture_path("types_mixed"), data_only=True)
    try:
        ws = wb["Sheet1"]
        assert ws["A1"].value == "id"
        # Column B holds digit strings stored as text -> numeric_text.
        assert ws["B2"].value == "007"
        assert isinstance(ws["B2"].value, str)
        # Column C holds real datetimes.
        assert hasattr(ws["C2"].value, "year")
        # Column D interleaves ints and strings (mixed).
        assert ws["D2"].value == 100
        assert ws["D3"].value == "N/A"
    finally:
        wb.close()


def test_left_margin_cols_table_offset(fixture_path) -> None:
    """left_margin_cols: filler column A; table occupies columns C-E."""

    wb = openpyxl.load_workbook(
        fixture_path("left_margin_cols"), data_only=True
    )
    try:
        ws = wb["Sheet1"]
        assert ws["A1"].value == "참고사항"  # description text, not the table
        assert ws["C1"].value == "sku"  # table header starts at column C (3)
        assert ws["E1"].value == "price"  # right boundary column E (5)
        assert ws["C2"].value == "A-1"  # first data row
        assert ws.max_column == 5
    finally:
        wb.close()


def test_mixed_sheets_has_readme_and_data(fixture_path) -> None:
    """mixed_sheets: non-tabular 'README' sheet + tabular 'Data' sheet."""

    wb = openpyxl.load_workbook(fixture_path("mixed_sheets"), data_only=True)
    try:
        assert wb.sheetnames == ["README", "Data"]
        readme = wb["README"]
        assert readme.max_column == 1  # single sparse text column
        data = wb["Data"]
        assert data["A1"].value == "item"
        assert data.max_column == 3
    finally:
        wb.close()


def test_wide_sparse_timeseries_layout(fixture_path) -> None:
    """wide_sparse_timeseries: dense Period header, sparse date-axis body."""

    wb = openpyxl.load_workbook(
        fixture_path("wide_sparse_timeseries"), data_only=True
    )
    try:
        ws = wb["Quarterly Series"]
        assert ws.max_column == 12
        assert ws.max_row == 16
        assert ws["A4"].value == "Period"
        assert ws["B4"].value == "Q:TS:001"
        assert hasattr(ws["A5"].value, "year")
        assert ws["B5"].value == 100
        assert ws["C5"].value is None
    finally:
        wb.close()


def test_hidden_sheet_states(fixture_path) -> None:
    """hidden_sheet: visible/hidden/veryHidden sheet states (issue #6)."""

    wb = openpyxl.load_workbook(fixture_path("hidden_sheet"), data_only=True)
    try:
        assert wb.sheetnames == ["Visible", "Hidden", "VeryHidden"]
        assert wb["Visible"].sheet_state == "visible"
        assert wb["Hidden"].sheet_state == "hidden"
        assert wb["VeryHidden"].sheet_state == "veryHidden"
    finally:
        wb.close()


def test_blank_run_terminates_layout(fixture_path) -> None:
    """blank_run_terminates: data 2-5, blank run 6-7, noise 9-10 (issue #7)."""

    wb = openpyxl.load_workbook(
        fixture_path("blank_run_terminates"), data_only=True
    )
    try:
        ws = wb["Sheet1"]
        assert ws["A1"].value == "name"  # header (1-based row 1)
        assert ws["A5"].value == "A-4"  # last data row (data_end_row=5)
        # Rows 6-7 are the blank-run terminator.
        assert ws["A6"].value is None
        assert ws["A7"].value is None
        # Noise block beyond the terminator.
        assert ws["A9"].value == "기타 메모"
        assert ws["A10"].value == "Z-9"
        assert ws.max_row == 10
        assert ws.max_column == 3
    finally:
        wb.close()


def test_interior_blank_layout(fixture_path) -> None:
    """interior_blank: data 2-3, single blank row 4, data 5-6 (MEDIUM #4)."""

    wb = openpyxl.load_workbook(
        fixture_path("interior_blank"), data_only=True
    )
    try:
        ws = wb["Sheet1"]
        assert ws["A1"].value == "name"  # header row 1
        assert ws["A2"].value == "A-1"  # first data row
        assert ws["A3"].value == "A-2"
        assert ws["A4"].value is None  # single interior blank row 4
        assert ws["A5"].value == "A-3"
        assert ws["A6"].value == "A-4"  # last data row
        assert ws.max_row == 6
        assert ws.max_column == 3
    finally:
        wb.close()


def test_empty_sheet_is_empty(fixture_path) -> None:
    """empty_sheet: no cells written; openpyxl reports a 1x1 used range."""

    wb = openpyxl.load_workbook(fixture_path("empty_sheet"), data_only=True)
    try:
        ws = wb["Sheet1"]
        assert ws["A1"].value is None
        assert ws.max_row == 1
        assert ws.max_column == 1
    finally:
        wb.close()


def test_header_only_has_no_data_rows(fixture_path) -> None:
    """header_only: a header row 1 with nothing beneath it."""

    wb = openpyxl.load_workbook(fixture_path("header_only"), data_only=True)
    try:
        ws = wb["Sheet1"]
        assert ws["A1"].value == "a"
        assert ws.max_row == 1  # header only, no data
        assert ws["A2"].value is None
    finally:
        wb.close()


def test_no_header_is_pure_data(fixture_path) -> None:
    """no_header: homogeneous data from row 1, no string header row."""

    wb = openpyxl.load_workbook(fixture_path("no_header"), data_only=True)
    try:
        ws = wb["Sheet1"]
        assert ws["A1"].value == 1  # row 1 is data, not a string header
        assert ws.max_row == 5
        assert ws.max_column == 3
    finally:
        wb.close()


def test_corrupt_fixture_cannot_be_opened(fixture_path) -> None:
    """corrupt: truncated zip -> openpyxl raises BadZipFile; PK magic, not OLE."""

    path = fixture_path("corrupt")
    with pytest.raises(zipfile.BadZipFile):
        openpyxl.load_workbook(path)

    head = path.read_bytes()[:8]
    assert head.startswith(_ZIP_LOCAL_MAGIC)
    assert not olefile.isOleFile(str(path))


def test_encrypted_fixture_cannot_be_opened(fixture_path) -> None:
    """encrypted: real password-protected .xlsx; OLE2 magic distinguishes it."""

    path = fixture_path("encrypted")
    with pytest.raises(zipfile.BadZipFile):
        openpyxl.load_workbook(path)

    head = path.read_bytes()[:8]
    assert head == _OLE2_MAGIC
    assert olefile.isOleFile(str(path))


def test_encrypted_fixture_is_decryptable_with_password(fixture_path) -> None:
    """encrypted: msoffcrypto confirms encryption and the known password works."""

    msoffcrypto = pytest.importorskip("msoffcrypto")
    path = fixture_path("encrypted")
    with path.open("rb") as handle:
        office = msoffcrypto.OfficeFile(handle)
        assert office.is_encrypted()
        office.load_key(password=ENCRYPTED_PASSWORD)  # must not raise
