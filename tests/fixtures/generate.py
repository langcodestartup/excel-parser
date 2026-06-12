"""Deterministic fixture-corpus generator (implementation plan §5.1).

Every sample workbook in the spec §5.1 corpus is synthesized *programmatically*
with openpyxl so the corpus is reproducible and free of hand-editing. All cell
data is fixed (deterministic) so downstream golden/coordinate tests can pin
exact positions.

Coordinate convention in this module and in the companion ``FIXTURES`` table:
all row/column numbers in docstrings and metadata are **openpyxl 1-based**
(the inspection / spec domain, [D1]). The Plan Aggregator is the only place
that converts to pandas 0-based; fixtures never speak pandas coordinates.

The encrypted sample is produced with ``msoffcrypto-tool`` (a genuine
OLE2/CFB-wrapped, password-protected ``.xlsx``); the corrupt sample is a
truncated zip. Both currently raise :class:`zipfile.BadZipFile` from openpyxl,
but they are distinguishable by their leading magic bytes (encrypted = OLE2
``d0cf11e0``, corrupt = ``PK\\x03\\x04`` of a *truncated* zip), which the
Phase 1 loader uses to map them to :class:`EncryptedWorkbookError` vs
:class:`CorruptWorkbookError`.

Run as a module to (re)generate the full corpus into this directory::

    python -m tests.fixtures.generate            # -> tests/fixtures/*.xlsx
    python tests/fixtures/generate.py /some/dir   # -> /some/dir/*.xlsx
"""

from __future__ import annotations

import datetime as _dt
import io
import re
import sys
import zipfile
from collections.abc import Callable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

#: Fixed timestamp stamped into every workbook's document properties so the
#: saved bytes are reproducible (no wall-clock created/modified leakage).
_FIXED_PROPERTY_TS = _dt.datetime(2020, 1, 1, 0, 0, 0)

#: The same fixed timestamp rendered as W3CDTF (the format openpyxl writes into
#: ``docProps/core.xml`` for ``dcterms:created`` / ``dcterms:modified``). We pin
#: the *XML payload* to this string because openpyxl overwrites the
#: ``dcterms:modified`` element with the wall-clock save time regardless of
#: ``wb.properties.modified`` (HIGH #1).
_FIXED_PROPERTY_TS_W3C = _FIXED_PROPERTY_TS.strftime("%Y-%m-%dT%H:%M:%SZ")

#: The OOXML core-properties part whose ``dcterms:*`` timestamps must be pinned.
_CORE_PROPS_MEMBER = "docProps/core.xml"

#: Matches a ``<dcterms:created ...>...</dcterms:created>`` (or ``modified``)
#: element so its inner W3CDTF text can be replaced with the fixed timestamp.
_DCTERMS_TS_RE = re.compile(
    rb"(<dcterms:(?:created|modified)\b[^>]*>)[^<]*(</dcterms:(?:created|modified)>)"
)

#: The minimum date_time the zip format can encode (1980-01-01 00:00:00).
#: Every archive member is normalized to this so the container bytes are stable.
_ZIP_EPOCH = (1980, 1, 1, 0, 0, 0)

# ---------------------------------------------------------------------------
# Fixture metadata (single source of truth for tests; 1-based coordinates)
# ---------------------------------------------------------------------------


class FixtureSpec:
    """Static metadata describing one generated fixture (1-based coords).

    Attributes:
        filename: Output ``.xlsx`` file name.
        openable: Whether openpyxl can open the file (False for
            corrupt/encrypted negatives).
        description: Human-readable description of the structure, with the
            salient 1-based coordinates (header row, subtotal/total rows,
            column boundaries) documented for downstream tests.
    """

    def __init__(self, filename: str, openable: bool, description: str) -> None:
        self.filename = filename
        self.openable = openable
        self.description = description


#: Canonical corpus metadata. Keys are stable fixture identifiers used by
#: tests; values carry the file name, openability, and a coordinate summary.
FIXTURES: dict[str, FixtureSpec] = {
    "header_simple": FixtureSpec(
        "header_simple.xlsx",
        True,
        "Sheet 'Sheet1'. Header at row 1 (A1:D1). Data rows 2-6 (5 rows). "
        "Columns A-D used. No subtotals.",
    ),
    "header_offset": FixtureSpec(
        "header_offset.xlsx",
        True,
        "Sheet 'Sheet1'. Title block rows 1-3 (sparse). Header at row 4 "
        "(A4:D4). Data rows 5-9 (5 rows). Columns A-D.",
    ),
    "offset_plus_subtotals": FixtureSpec(
        "offset_plus_subtotals.xlsx",
        True,
        "Sheet 'Sheet1'. [D1] coordinate-regression fixture. Title rows 1-3. "
        "Header at row 4 (A4:D4). Data rows 5-7, subtotal row 8 ('소계'), "
        "data rows 9-11, subtotal row 12 ('소계'), grand-total row 13 "
        "('합계'). skip_rows (1-based) = [8, 12, 13]. data_start_row=5, "
        "data_end_row=11. Columns A-D.",
    ),
    "merged_header": FixtureSpec(
        "merged_header.xlsx",
        True,
        "Sheet 'Sheet1'. Header row 1 with a single-level merge A1:B1 "
        "('이름') spanning two leaf columns; C1='점수'. The merge leaves B1 "
        "empty, so the header row reads ('이름', None, '점수') — a populated "
        "lead cell, an interior gap, then another populated cell. Phase 11a "
        "behavior (plan v2 Task 11.1): the Merge Scanner collects A1:B1 "
        "before boundary detection, the empty B1 is virtually filled, and the "
        "bridged full-width span A-C resolves header_row=1, data_start_row=2, "
        "data_end_row=5, data_left_col/data_right_col=None (full width -> no "
        "usecols). The body merge A6:A7 (rows 6-7, '그룹'/'정대만'/'송태섭' "
        "demo) is a separate kind='body' block and is NOT part of the table "
        "body: a fully merge-grouped row group trailing the merge-free body "
        "rows 2-5 is excluded (data_end_row is 5, NEVER 7 — plan v2 §5 "
        "fixture-contradiction guard) with a visible 'trailing merged-row "
        "group (rows 6-7)' warning, while the A6:A7 forward-fill note still "
        "reaches ReadPlan.notes. Pre-11a (no Merge Scanner, e.g. a standalone "
        "BoundaryDetector run without collected spans) the §7.2 / MEDIUM #7 "
        "deferral applies instead: all boundaries UNRESOLVED (None) plus a "
        "'discarded pending merge analysis' warning. Columns A-C.",
    ),
    "multi_level_header": FixtureSpec(
        "multi_level_header.xlsx",
        True,
        "Sheet 'Sheet1'. Two-level header. Row 1 group "
        "merges: A1:B1 ('상반기'), C1:D1 ('하반기'). Row 2 leaf headers "
        "A2..D2 ('1월','2월','3월','4월'). Data rows 3-6 (4 rows). Columns "
        "A-D. header_row resolves to 2; both row-1 merges classify as "
        "kind='header'; is_multi_level_header=True. data_start_row=3, "
        "data_end_row=6, data_left/right_col=None (full width -> no "
        "usecols), dtype_map={} (all columns number). Phase 11b expected "
        "plan (plan v2 Task 11.2): header=[0, 1] (0-based absolute band "
        "rows 1-2), skiprows=[] (leading rows are NOT absorbed — pandas "
        "merged-cell header fill happens at pre-skip absolute rows), "
        "nrows=4. Loaded MultiIndex columns flatten to "
        "['상반기 / 1월','상반기 / 2월','하반기 / 3월','하반기 / 4월'] "
        "(pandas forward-fills the merged group labels). Per-column data "
        "sums: 1월=46, 2월=86, 3월=126, 4월=166.",
    ),
    "multi_level_numeric_text": FixtureSpec(
        "multi_level_numeric_text.xlsx",
        True,
        "Sheet 'Sheet1'. [Phase 11b dtype x MultiIndex golden] Two-level "
        "header over a numeric_text leaf column. Row 1 group merges: A1:B1 "
        "('기본'), C1:D1 ('실적'). Row 2 leaf headers A2..D2 "
        "('코드','수량','단가','금액'). Data rows 3-5 (3 rows): 코드 = "
        "'007'/'012'/'034' digit strings stored as text (numeric_text); "
        "수량/단가/금액 numbers. Columns A-D. Expected (1-based): "
        "header_row=2 (heuristic), is_multi_level_header=True, "
        "data_start_row=3, data_end_row=5, data_left/right_col=None (full "
        "width -> no usecols), dtype_map={'0': 'string'} ([D5] positional "
        "key). Phase 11b expected plan: header=[0, 1], skiprows=[], "
        "nrows=3. Loaded columns flatten to ['기본 / 코드','기본 / 수량',"
        "'실적 / 단가','실적 / 금액'] and the 코드 leading zeros survive "
        "('007'), proving the positional dtype_map key stays position-valid "
        "under header=list (pandas 3.0.3, measured).",
    ),
    "titled_multi_level": FixtureSpec(
        "titled_multi_level.xlsx",
        True,
        "Sheet 'Sheet1'. [Issue #7 regression] A merged title banner A1:H1 "
        "('2026년 상반기 지역별 매출 실적 보고서') above a blank row 2 and a "
        "two-level header band rows 3-4: vertical merges A3:A4 ('지역') and "
        "B3:B4 ('제품코드'), group merges C3:E3 ('1분기') and F3:H3 "
        "('2분기'), leaf labels C4..H4 ('1월'..'6월'). Data rows 5-6, "
        "columns A-H; column B holds digit strings '00123'/'00789' "
        "(numeric_text, leading zeros must survive). header_row resolves to "
        "4 and every merge with min_row <= 4 classifies kind='header' — "
        "including the title — so the merged-header row set {1, 3} is "
        "non-contiguous. Expected plan (issue #7 fix): the maximal "
        "contiguous run ending at the leaf (rows 3-4) survives as "
        "header=[2, 3] (0-based absolute [D1]); the disconnected title row "
        "[1] is excluded from the band with a visible warning (spec §8 no "
        "silent loss) instead of vetoing the whole band into leaf-only "
        "'Unnamed: N' loading. Loaded columns flatten to ['지역','제품코드',"
        "'1분기 / 1월','1분기 / 2월','1분기 / 3월','2분기 / 4월',"
        "'2분기 / 5월','2분기 / 6월'] (vertical-merge anchors fill the "
        "group level; empty leaf cells drop from the join), and records "
        "keys == columns[].resolved_name positionally.",
    ),
    "stacked_multi_level": FixtureSpec(
        "stacked_multi_level.xlsx",
        True,
        "Sheet 'Sheet1'. [Adversarial review MEDIUM #1 regression] A flat "
        "table stacked ABOVE a two-level-header table. Table 1: header row 1 "
        "(A1:D1 = '부서','인원','예산','비고'), data rows 2-4 (3 rows), "
        "columns A-D. Rows 5-6 fully blank (BLANK_RUN separator). Table 2: "
        "row 7 group merges A7:B7 ('상반기') and C7:D7 ('하반기'), row 8 "
        "leaf headers A8..D8 ('1월','2월','3월','4월'), data rows 9-11 "
        "(3 rows), columns A-D. Row bands (1-based, inclusive): [1..4] and "
        "[7..11]. max_row=11, max_col=4. The sheet-level Merge Analyzer "
        "classifies the row-7 group merges against the MIRRORED sheet header "
        "(row 1, blocks[0]) as kind='body'; the aggregator's band-scoped "
        "path must RE-classify them against T2's OWN header (row 8) as "
        "kind='header' (review MEDIUM #1), so T2 promotes to a multi-level "
        "header. Expected T2 plan (0-based [D1]): header=[6, 7] (absolute "
        "band rows 7-8), skiprows=[] (leading rows NOT absorbed), nrows=3, "
        "no usecols (full width), dtype_map={} (all columns number), and NO "
        "'body merge' forward-fill note on either block. Loaded T2 columns "
        "flatten to ['상반기 / 1월','상반기 / 2월','하반기 / 3월',"
        "'하반기 / 4월'] with per-column sums 33/63/93/123; T1 stays "
        "['부서','인원','예산','비고'] with 3 rows.",
    ),
    "types_mixed": FixtureSpec(
        "types_mixed.xlsx",
        True,
        "Sheet 'Sheet1'. Header row 1: A='id'(number), B='code'(numeric_text, "
        "digit strings stored as text), C='date'(date cells), D='mixed'(int "
        "and string interleaved). Data rows 2-7 (6 rows). Columns A-D.",
    ),
    "left_margin_cols": FixtureSpec(
        "left_margin_cols.xlsx",
        True,
        "Sheet 'Sheet1'. Left description column A is filler text only "
        "(rows 1-7). The actual table occupies columns C-E: header row 1 "
        "(C1:E1), data rows 2-7. data_left_col=3 (C), data_right_col=5 (E); "
        "usecols expected 'C:E'.",
    ),
    "left_margin_with_subtotal": FixtureSpec(
        "left_margin_with_subtotal.xlsx",
        True,
        "Sheet 'Sheet1'. [Phase 13 Step 3, L7 keyword-coordinate regression] "
        "Variant of left_margin_cols whose left margin is NON-empty on the "
        "subtotal row: column A carries filler notes at rows 1-3 and row 5 "
        "(A5='중간 점검용 행' — deliberately NOT a skip keyword). The table "
        "occupies columns C-E: header row 1 (C1:E1='sku','qty','price'), "
        "data rows 2-4, subtotal row 5 (C5='소계', D5 empty, E5=7.5 — span "
        "density 2/3, ABOVE the low-density threshold, so only the keyword "
        "rule can catch it), data rows 6-7. Expected (1-based): "
        "header_row=1, data_left_col=3 (C), data_right_col=5 (E), usecols "
        "'C:E', data_start_row=2, data_end_row=7, skip_rows=[5]. A sheet-"
        "column-A keyword scan reads the row-5 margin note instead of "
        "'소계' and silently leaks the subtotal into the data (the plan v2 "
        "§3 review-checklist trap: left_margin_cols' margin is empty on "
        "every data row, so an A-column implementation stays green there); "
        "the data_left_col-anchored scan must flag it. Loaded frame: 5 "
        "rows, qty sum 150, no '소계' anywhere.",
    ),
    "mixed_sheets": FixtureSpec(
        "mixed_sheets.xlsx",
        True,
        "Workbook with two sheets. Sheet 'README' is a non-tabular "
        "description sheet (single sparse text column, "
        "is_tabular_candidate expected False). Sheet 'Data' is a normal "
        "table: header row 1 (A1:C1), data rows 2-5, columns A-C.",
    ),
    "cover_offset": FixtureSpec(
        "cover_offset.xlsx",
        True,
        "Single-sheet cover '표지' whose sparse text starts in column B "
        "(B2/B4/B6). max_col=2 fooled the legacy max_col>1 gate, but only one "
        "column is populated -> is_tabular_candidate expected False (issue #3).",
    ),
    "cover_sparse": FixtureSpec(
        "cover_sparse.xlsx",
        True,
        "Single-sheet cover '표지' with scattered cells across 3 columns "
        "(B2/E4/C6): populated_cols=3, populated_rows=3, filled=3, "
        "density=0.333 -> is_tabular_candidate expected False via the density "
        "rule (issue #3).",
    ),
    "sparse_real_table": FixtureSpec(
        "sparse_real_table.xlsx",
        True,
        "Genuine 4-column table with many missing cells: header row + 5 data "
        "rows, populated_cols=4, populated_rows=6, filled=14, density=0.583 -> "
        "ABOVE NON_TABULAR_DENSITY_THRESHOLD (0.5) so is_tabular_candidate "
        "expected True. Pins the density-rule margin so the threshold cannot "
        "be raised past 0.583 without a red test (issue #3).",
    ),
    "hidden_sheet": FixtureSpec(
        "hidden_sheet.xlsx",
        True,
        "Workbook with three sheets in order: 'Visible' "
        "(sheet_state='visible', is_visible expected True; header row 1 "
        "A1:C1, data rows 2-4, columns A-C), 'Hidden' "
        "(sheet_state='hidden', is_visible expected False; header row 1 "
        "A1:B1, data rows 2-3), and 'VeryHidden' "
        "(sheet_state='veryHidden', is_visible expected False; header row 1 "
        "A1:B1, data rows 2-3).",
    ),
    "blank_run_terminates": FixtureSpec(
        "blank_run_terminates.xlsx",
        True,
        "Sheet 'Sheet1'. [Phase 3 Boundary] header row 1 (A1:C1), data rows "
        "2-5 (4 rows), then a BLANK_RUN of 2 fully-empty rows (rows 6-7) that "
        "terminates the table, followed by an unrelated noise block: a stray "
        "label at row 9 (A9='기타 메모') and a trailing noise data row at row "
        "10 (A10:C10). Expected (1-based): data_start_row=2, data_end_row=5; "
        "the rows 6-7 blank run is the terminator and rows 9-10 lie beyond it "
        "(not part of the table). skip_rows expected empty (no interior "
        "subtotal). max_row=10, max_col=3. Phase 10b expected behavior (W-A "
        "review LOW #9, recalibrated by issue #8): row bands are [1..5] and "
        "[9..10]. Pre-issue-#8 the noise band crossed the 0.5 threshold only "
        "via the free type-consistency of its single-row lookahead window and "
        "was extracted as 'Sheet1!T2' with the 1-column verify advisory "
        "(review LOW #7); with the §7.1 lookahead-evidence factor it scores "
        "below threshold and is REJECTED — exactly one table 'Sheet1!T1', and "
        "the not-a-table judgment for rows 9-10 surfaces as a warning (spec "
        "§8). The LOW #7 advisory contract stays pinned at the plan level "
        "(test_multi_table). T1 and the flat sheet fields keep the exact v1 "
        "golden values above.",
    ),
    "interior_blank": FixtureSpec(
        "interior_blank.xlsx",
        True,
        "Sheet 'Sheet1'. [Phase 3 Boundary / MEDIUM #4] header row 1 (A1:C1), "
        "data rows 2-3, a SINGLE interior blank row 4 (below the BLANK_RUN of "
        "2 threshold), then data rows 5-6. The lone blank row 4 must be "
        "recorded in skip_rows so it never leaks into the loaded frame as an "
        "all-NaN row. Expected (1-based): data_start_row=2, data_end_row=6, "
        "skip_rows=[4]. max_row=6, max_col=3.",
    ),
    "empty_sheet": FixtureSpec(
        "empty_sheet.xlsx",
        True,
        "Sheet 'Sheet1' is completely empty (no cells written). max_row/"
        "max_col effectively 1/1 in openpyxl; data_start_row expected None.",
    ),
    "header_only": FixtureSpec(
        "header_only.xlsx",
        True,
        "Sheet 'Sheet1'. Header row 1 (A1:C1) with zero data rows below. "
        "data_start_row expected None; ReadPlan header-only.",
    ),
    "no_header": FixtureSpec(
        "no_header.xlsx",
        True,
        "Sheet 'Sheet1'. Pure data, no header: rows 1-5 are all homogeneous "
        "data (each row: int, int, float). Header heuristic expected to "
        "fail -> needs_manual_header. Columns A-C.",
    ),
    "large_table": FixtureSpec(
        "large_table.xlsx",
        True,
        "Sheet 'Sheet1'. [Phase 8 performance/sampling] header row 1 (A1:D1: "
        "'id','name','amount','flag'), LARGE_TABLE_DATA_ROWS (5000) data rows "
        "2-5001, then a trailing grand-total row 5002 ('합계'). Used to prove "
        "inspect() streams/samples rather than fully scanning: the Type "
        "Profiler samples at most TYPE_SAMPLE_ROWS (200) of the 5000 data rows, "
        "the Header Locator reads only the top HEADER_SCAN_ROWS (20), and no "
        "analyzer materializes the whole sheet. Expected (1-based): "
        "header_row=1, data_start_row=2, data_end_row=5001, skip_rows=[5002] "
        "(the trailing total). max_row=5002, max_col=4.",
    ),
    "multi_table_stacked": FixtureSpec(
        "multi_table_stacked.xlsx",
        True,
        "Sheet 'Sheet1'. [Phase 10 multi-table] Two vertically stacked tables "
        "separated by a BLANK_RUN (2) of empty rows. Table 1: header row 1 "
        "(A1:C1 = '부서','인원','예산'), data rows 2-4 (3 rows), columns A-C. "
        "Blank separator rows 5-6 (fully empty). Table 2: header row 7 "
        "(A7:D7 = '제품명','단가','재고','비고'), data rows 8-10 (3 rows), "
        "columns A-D. Row bands (1-based, inclusive): [1..4] and [7..10]. "
        "max_row=10, max_col=4. Phase 10a: the second band carries a header "
        "candidate, so inspect()/extract() must record a multi-block warning "
        "(v1 still extracts a single block). Phase 10b golden: T1 columns "
        "['부서','인원','예산'] with 3 rows, T2 columns "
        "['제품명','단가','재고','비고'] with 3 rows.",
    ),
    "stacked_uneven_width": FixtureSpec(
        "stacked_uneven_width.xlsx",
        True,
        "Sheet 'Sheet1'. [Phase 10 guard #1: score-denominator dilution] Two "
        "stacked tables of UNEVEN width separated by a BLANK_RUN (2) of empty "
        "rows. Table 1 (narrow, 3 cols): header row 1 (A1:C1 = "
        "'코드','명칭','수량'), data rows 2-4 (3 rows), columns A-C. Blank "
        "separator rows 5-6. Table 2 (wide, 8 cols): header row 7 (A7:H7 = "
        "'일자','지점','담당','품목','단가','수량','금액','비고'), data rows "
        "8-10 (3 rows), columns A-H. Row bands (1-based, inclusive): [1..4] "
        "and [7..10]. max_row=10, max_col=8. A sheet-global max_col (8) "
        "denominator dilutes the narrow 3-column table's header/density "
        "scores; band-local column counts must be used instead (plan v2 §4 "
        "Task 10.2 guard 1).",
    ),
    "title_between_tables": FixtureSpec(
        "title_between_tables.xlsx",
        True,
        "Sheet 'Sheet1'. [W-A review HIGH regression] A leading 1-row title "
        "band, two stacked tables, and a trailing 1-row string footnote band, "
        "each separated by a BLANK_RUN (2) of empty rows. Row 1: A1='2026년 "
        "1분기 부서별 집계' (title; a 1-row all-string band scores exactly "
        "0.500 — AT the default header threshold — but resolves no data, so "
        "it must be judged not a table). Rows 2-3 blank. Table 1: header row "
        "4 (A4:C4 = '부서','인원','예산'), data rows 5-7 (3 rows), columns "
        "A-C. Rows 8-9 blank. Table 2: header row 10 (A10:C10 = '품목','수량',"
        "'금액'), data rows 11-13 (3 rows), columns A-C. Rows 14-15 blank. "
        "Row 16: A16='주: 단위는 천원' (string footnote band — same "
        "not-a-table judgment). Row bands (1-based, inclusive): [1..1], "
        "[4..7], [10..13], [16..16]. max_row=16, max_col=3. Expected: exactly "
        "two tables — 'Sheet1!T1' header_row=4 / data 5-7 and 'Sheet1!T2' "
        "header_row=10 / data 11-13 — with the flat mirror following blocks[0] "
        "(header_row=4, the v1 result); the title/footnote bands enter no "
        "block (warnings only), every lower-table value appears in exactly "
        "one TableResult, and no extracted frame contains an all-NaN row.",
    ),
    "formulas": FixtureSpec(
        "formulas.xlsx",
        True,
        "Sheet 'Sheet1'. [Phase 12 formula detection golden] Header row 1 "
        "(A1:D1 = '품목','수량','단가','금액'). Data rows 2-5 (4 rows), "
        "columns A-D. Column D cells are formulas '=B{r}*C{r}' written by "
        "openpyxl, which stores an EMPTY cached value (<v/>) — in data mode "
        "(data_only=True) column D reads all None, while formula mode "
        "(data_only=False) reads the formula strings. Body values: "
        "수량 B=[2,3,4,5], 단가 C=[100,200,300,400] (D would evaluate to "
        "[200,600,1200,2000] but no cache exists, so this is never readable "
        "from data mode). Expected (1-based): header_row=1, data_start_row=2, "
        "data_end_row=5, data_left/right_col=None (full width). Phase 12 "
        "expected: columns[3].has_formula=True, read_hint='as_formula', a "
        "'formula cache empty (file never opened in Excel?)' warning, "
        "dtype_map WITHOUT key '3' (as_formula columns skip dtype inference; "
        "'0' stays 'string' for the 품목 text column) and an as_formula "
        "advisory note in ReadPlan.notes. Columns A-C: has_formula=False, "
        "read_hint='as_value'.",
    ),
    "formulas_cached": FixtureSpec(
        "formulas_cached.xlsx",
        True,
        "Sheet 'Sheet1'. [Phase 12 probe 2: as_value branch] Identical layout "
        "to 'formulas' (header row 1 A1:D1, data rows 2-5, D column "
        "'=B{r}*C{r}' formulas), but cached results are injected into the "
        "sheet XML by zip post-processing (FORMULA_CACHE_VALUES: D2=200, "
        "D3=600, D4=1200, D5=2000 — what Excel would cache on a recalc "
        "save; openpyxl alone never writes caches). Data mode therefore "
        "reads real numbers for D while formula mode still reads the "
        "formula strings. Phase 12 expected: columns[3].has_formula=True, "
        "read_hint='as_value', NO cache-empty warning, no as_formula note; "
        "column D profiles as number so dtype_map has no key '3' anyway. "
        "Loaded D values [200, 600, 1200, 2000] (sum 4000).",
    ),
    "corrupt": FixtureSpec(
        "corrupt.xlsx",
        False,
        "Truncated-zip bytes (first half of a valid .xlsx). Leading bytes "
        "'PK\\x03\\x04'. openpyxl raises zipfile.BadZipFile; loader maps to "
        "CorruptWorkbookError.",
    ),
    "encrypted": FixtureSpec(
        "encrypted.xlsx",
        False,
        "Genuine password-protected .xlsx (password 'secret') produced via "
        "msoffcrypto-tool. OLE2/CFB container (leading bytes 'd0cf11e0'). "
        "openpyxl raises zipfile.BadZipFile; loader distinguishes via the "
        "OLE2 magic / msoffcrypto and maps to EncryptedWorkbookError.",
    ),
}

#: Password used for the encrypted fixture (also exported for loader tests).
ENCRYPTED_PASSWORD = "secret"

#: Number of data rows in the ``large_table`` performance fixture (spec §8 /
#: Phase 8). Large enough that a full materialization of every row would be a
#: meaningful memory/time cost, so the streaming + sampling guarantees are
#: observable; the trailing total row sits at ``2 + LARGE_TABLE_DATA_ROWS``.
LARGE_TABLE_DATA_ROWS = 5000

#: Number of data rows in the on-demand 100k-row performance-smoke workbook
#: (plan v2 Phase 13 Step 4). The builder (:func:`build_perf_100k`) is
#: deliberately **excluded** from ``FIXTURES``/``BUILDERS`` so the default
#: corpus (and every corpus-parametrized test) never pays the build cost; only
#: the ``@pytest.mark.slow`` memory-smoke test materializes it, into a pytest
#: tmp dir (see ``tests/conftest.py``).
PERF_TABLE_DATA_ROWS = 100_000


# ---------------------------------------------------------------------------
# Per-fixture builders (each returns the in-memory bytes of one .xlsx)
# ---------------------------------------------------------------------------


def _pin_core_props(payload: bytes) -> bytes:
    """Pin the ``dcterms:created`` / ``dcterms:modified`` text in core.xml.

    openpyxl overwrites ``dcterms:modified`` (and may re-render ``created``)
    with the *wall-clock* save time in ``docProps/core.xml`` regardless of
    ``wb.properties.modified``. Normalizing only the zip-member ``date_time``
    therefore leaves a wall-clock timestamp inside the XML payload, making the
    bytes non-deterministic across a save that crosses a second boundary. We
    rewrite the inner W3CDTF text of both elements to the fixed timestamp so the
    payload is stable (HIGH #1).

    Args:
        payload: The raw bytes of ``docProps/core.xml``.

    Returns:
        The payload with all ``dcterms:created``/``dcterms:modified`` inner text
        replaced by :data:`_FIXED_PROPERTY_TS_W3C`.
    """

    replacement = rb"\g<1>" + _FIXED_PROPERTY_TS_W3C.encode("ascii") + rb"\g<2>"
    return _DCTERMS_TS_RE.sub(replacement, payload)


def _normalize_zip_timestamps(data: bytes) -> bytes:
    """Rewrite a zip so every member is byte-stable (timestamps + core props).

    openpyxl writes each archive member with the wall-clock time of the save
    (in the zip member ``date_time``) *and* stamps the wall-clock save time into
    ``docProps/core.xml``'s ``dcterms:modified`` element, both of which make the
    ``.xlsx`` bytes non-deterministic. We repack the archive member-by-member,
    forcing every :class:`zipfile.ZipInfo.date_time` to ``_ZIP_EPOCH`` and a
    fixed compression, *and* rewriting the ``docProps/core.xml`` payload's
    ``dcterms:*`` timestamps to the fixed value (HIGH #1), so two saves of the
    same logical workbook are byte-identical.

    Args:
        data: The original ``.xlsx`` (zip) bytes.

    Returns:
        The repacked, fully timestamp-normalized zip bytes.
    """

    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data)) as src:
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
            for info in src.infolist():
                payload = src.read(info.filename)
                if info.filename == _CORE_PROPS_MEMBER:
                    payload = _pin_core_props(payload)
                new_info = zipfile.ZipInfo(
                    filename=info.filename, date_time=_ZIP_EPOCH
                )
                new_info.compress_type = zipfile.ZIP_DEFLATED
                new_info.external_attr = info.external_attr
                new_info.internal_attr = info.internal_attr
                new_info.create_system = info.create_system
                dst.writestr(new_info, payload)
    return out.getvalue()


def _save_bytes(wb: Workbook) -> bytes:
    """Serialize a workbook to byte-stable ``.xlsx`` bytes (deterministic).

    Both sources of non-determinism are pinned: the workbook's document
    ``created``/``modified`` properties are forced to ``_FIXED_PROPERTY_TS``,
    and the resulting zip's per-member timestamps are normalized to the zip
    epoch via :func:`_normalize_zip_timestamps`. Saving the same logical
    workbook twice therefore yields identical bytes.
    """

    wb.properties.created = _FIXED_PROPERTY_TS
    wb.properties.modified = _FIXED_PROPERTY_TS
    buf = io.BytesIO()
    wb.save(buf)
    return _normalize_zip_timestamps(buf.getvalue())


def _write_rows(
    ws: Worksheet, start_row: int, rows: list[list[object]]
) -> None:
    """Write ``rows`` into ``ws`` starting at 1-based ``start_row``, col A."""

    for offset, row in enumerate(rows):
        excel_row = start_row + offset
        for col_idx, value in enumerate(row, start=1):
            if value is None:
                continue
            ws.cell(row=excel_row, column=col_idx, value=value)


def build_header_simple() -> bytes:
    """Header at row 1, 5 data rows, columns A-D (see ``FIXTURES``)."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(
        ws,
        1,
        [
            ["name", "age", "city", "score"],
            ["Alice", 30, "Seoul", 90.5],
            ["Bob", 25, "Busan", 80.0],
            ["Carol", 41, "Incheon", 75.5],
            ["Dave", 38, "Daegu", 88.0],
            ["Eve", 29, "Gwangju", 95.5],
        ],
    )
    return _save_bytes(wb)


def build_header_offset() -> bytes:
    """Title rows 1-3, header at row 4, 5 data rows (see ``FIXTURES``)."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "월간 판매 보고서"
    ws["A2"] = "작성일: 2026-01-01"
    ws["A3"] = "단위: 원"
    _write_rows(
        ws,
        4,
        [
            ["product", "region", "qty", "amount"],
            ["Widget", "North", 10, 1000],
            ["Gadget", "South", 5, 750],
            ["Widget", "East", 8, 800],
            ["Gizmo", "West", 12, 1440],
            ["Gadget", "North", 3, 450],
        ],
    )
    return _save_bytes(wb)


def build_offset_plus_subtotals() -> bytes:
    """[D1] regression: leading rows + mid subtotals + grand total.

    Title rows 1-3, header row 4, data rows 5-7 / 9-11, subtotal rows 8 & 12
    ('소계'), grand-total row 13 ('합계'). See ``FIXTURES`` for the exact
    1-based coordinates relied on by ``test_aggregator_coords``.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "부서별 비용 집계"
    ws["A2"] = "작성일: 2026-02-01"
    ws["A3"] = "단위: 천원"
    rows: list[list[object]] = [
        ["dept", "item", "month", "amount"],  # row 4 header
        ["영업", "교통비", 1, 100],  # row 5
        ["영업", "식대", 1, 200],  # row 6
        ["영업", "비품", 1, 50],  # row 7
        ["소계", None, None, 350],  # row 8 subtotal
        ["관리", "교통비", 1, 80],  # row 9
        ["관리", "식대", 1, 120],  # row 10
        ["관리", "비품", 1, 40],  # row 11
        ["소계", None, None, 240],  # row 12 subtotal
        ["합계", None, None, 590],  # row 13 grand total
    ]
    _write_rows(ws, 4, rows)
    return _save_bytes(wb)


def build_merged_header() -> bytes:
    """Single-level header merge A1:B1 plus a body merge A6:A7."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Header row 1: A1:B1 merged label, C1 leaf.
    ws["A1"] = "이름"
    ws["C1"] = "점수"
    ws.merge_cells("A1:B1")
    _write_rows(
        ws,
        2,
        [
            ["Kim", "김철수", 90],
            ["Lee", "이영희", 85],
            ["Park", "박민수", 70],
            ["Choi", "최지우", 60],
        ],
    )
    # Body merge demonstrating kind="body": group label spanning two rows.
    ws["A6"] = "그룹"
    ws.merge_cells("A6:A7")
    ws["B6"] = "정대만"
    ws["C6"] = 50
    ws["B7"] = "송태섭"
    ws["C7"] = 55
    return _save_bytes(wb)


def build_multi_level_header() -> bytes:
    """Two-level header (group merges row 1, leaf labels row 2). v1+ case."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "상반기"
    ws.merge_cells("A1:B1")
    ws["C1"] = "하반기"
    ws.merge_cells("C1:D1")
    _write_rows(
        ws,
        2,
        [
            ["1월", "2월", "3월", "4월"],  # row 2 leaf header
            [10, 20, 30, 40],  # row 3
            [11, 21, 31, 41],  # row 4
            [12, 22, 32, 42],  # row 5
            [13, 23, 33, 43],  # row 6
        ],
    )
    return _save_bytes(wb)


def build_multi_level_numeric_text() -> bytes:
    """Two-level header over a numeric_text leaf column (Phase 11b golden)."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "기본"
    ws.merge_cells("A1:B1")
    ws["C1"] = "실적"
    ws.merge_cells("C1:D1")
    _write_rows(
        ws,
        2,
        [
            ["코드", "수량", "단가", "금액"],  # row 2 leaf header
            ["007", 1, 1000, 1000],  # row 3
            ["012", 4, 500, 2000],  # row 4
            ["034", 7, 200, 1400],  # row 5
        ],
    )
    return _save_bytes(wb)


def build_titled_multi_level() -> bytes:
    """Merged title banner above a two-level header band (issue #7)."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "2026년 상반기 지역별 매출 실적 보고서"
    ws.merge_cells("A1:H1")
    ws["A3"] = "지역"
    ws.merge_cells("A3:A4")
    ws["B3"] = "제품코드"
    ws.merge_cells("B3:B4")
    ws["C3"] = "1분기"
    ws.merge_cells("C3:E3")
    ws["F3"] = "2분기"
    ws.merge_cells("F3:H3")
    _write_rows(
        ws,
        4,
        [
            # Row 4 leaves; A4/B4 stay covered by the vertical merges above.
            [None, None, "1월", "2월", "3월", "4월", "5월", "6월"],
            ["서울", "00123", 120, 135, 150, 142, 138, 160],  # row 5
            ["부산", "00789", 60, 65, 70, 72, 68, 75],  # row 6
        ],
    )
    return _save_bytes(wb)


def build_stacked_multi_level() -> bytes:
    """A flat table stacked above a two-level-header table (review MEDIUM #1).

    Table 1: header row 1 (A1:D1), data rows 2-4, columns A-D. Rows 5-6 blank
    (BLANK_RUN band separator). Table 2: group merges A7:B7 ('상반기') /
    C7:D7 ('하반기'), leaf header row 8, data rows 9-11. See ``FIXTURES`` for
    the expected band-scoped re-classification and the T2 multi-level plan.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(
        ws,
        1,
        [
            ["부서", "인원", "예산", "비고"],  # row 1 header (table 1)
            ["영업", 12, 3400, "본사"],  # row 2
            ["개발", 20, 5200, "본사"],  # row 3
            ["관리", 5, 1100, "지사"],  # row 4
            # rows 5-6 left fully empty -> BLANK_RUN band separator
        ],
    )
    ws["A7"] = "상반기"
    ws.merge_cells("A7:B7")
    ws["C7"] = "하반기"
    ws.merge_cells("C7:D7")
    _write_rows(
        ws,
        8,
        [
            ["1월", "2월", "3월", "4월"],  # row 8 leaf header (table 2)
            [10, 20, 30, 40],  # row 9
            [11, 21, 31, 41],  # row 10
            [12, 22, 32, 42],  # row 11
        ],
    )
    return _save_bytes(wb)


def build_types_mixed() -> bytes:
    """numeric_text / date / mixed columns (see ``FIXTURES``)."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "id"
    ws["B1"] = "code"
    ws["C1"] = "date"
    ws["D1"] = "mixed"
    # B is digit-strings stored as text (numeric_text); C is real dates;
    # D interleaves ints and strings (mixed).
    data = [
        (1, "007", _dt.date(2026, 1, 1), 100),
        (2, "012", _dt.date(2026, 1, 2), "N/A"),
        (3, "034", _dt.date(2026, 1, 3), 300),
        (4, "056", _dt.date(2026, 1, 4), "pending"),
        (5, "078", _dt.date(2026, 1, 5), 500),
        (6, "090", _dt.date(2026, 1, 6), 600),
    ]
    for offset, (a, b, c, d) in enumerate(data):
        r = 2 + offset
        ws.cell(row=r, column=1, value=a)
        # Force text storage for the numeric-looking code column.
        code_cell = ws.cell(row=r, column=2, value=b)
        code_cell.number_format = "@"
        ws.cell(row=r, column=3, value=c)
        ws.cell(row=r, column=4, value=d)
    return _save_bytes(wb)


def build_left_margin_cols() -> bytes:
    """Left filler column A; the real table occupies columns C-E."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Column A: free-form description text only (not part of the table).
    notes = [
        "참고사항",
        "이 표는 예시입니다",
        "열 C부터 표가 시작",
        "",
        "",
        "",
        "",
    ]
    for offset, note in enumerate(notes):
        if note:
            ws.cell(row=1 + offset, column=1, value=note)
    # Table in columns C(3)-E(5): header row 1, data rows 2-7.
    table = [
        ["sku", "qty", "price"],  # row 1 header
        ["A-1", 10, 1.5],
        ["A-2", 20, 2.5],
        ["A-3", 30, 3.5],
        ["A-4", 40, 4.5],
        ["A-5", 50, 5.5],
        ["A-6", 60, 6.5],
    ]
    for offset, row in enumerate(table):
        r = 1 + offset
        for col_offset, value in enumerate(row):
            ws.cell(row=r, column=3 + col_offset, value=value)
    return _save_bytes(wb)


def build_left_margin_with_subtotal() -> bytes:
    """Left filler column A (non-empty on the subtotal row) + C-E table (L7).

    The plan v2 §3 review-checklist trap made concrete: ``left_margin_cols``'
    margin is empty on every body row, so a keyword scan anchored at sheet
    column A still passes there. Here the margin note at A5 shadows the
    table's own '소계' label at C5 for an A-column implementation — only a
    ``data_left_col``-anchored scan flags row 5. See ``FIXTURES`` for the
    exact 1-based coordinates.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Column A margin notes; row 5 (the subtotal row) deliberately carries a
    # non-keyword note so an A-column keyword scan misses the subtotal.
    margin_notes = {
        1: "참고사항",
        2: "이 표는 예시입니다",
        3: "열 C부터 표가 시작",
        5: "중간 점검용 행",
    }
    for row, note in margin_notes.items():
        ws.cell(row=row, column=1, value=note)
    # Table in columns C(3)-E(5): header row 1, data rows 2-4 / 6-7, and a
    # subtotal row 5 whose span density (2/3) is above the low-density
    # threshold — only the keyword rule can catch it.
    table: list[list[object]] = [
        ["sku", "qty", "price"],  # row 1 header (C1:E1)
        ["A-1", 10, 1.5],  # row 2
        ["A-2", 20, 2.5],  # row 3
        ["A-3", 30, 3.5],  # row 4
        ["소계", None, 7.5],  # row 5 subtotal (C5 label, E5 price sum)
        ["B-1", 40, 4.5],  # row 6
        ["B-2", 50, 5.5],  # row 7
    ]
    for offset, row in enumerate(table):
        r = 1 + offset
        for col_offset, value in enumerate(row):
            if value is None:
                continue
            ws.cell(row=r, column=3 + col_offset, value=value)
    return _save_bytes(wb)


def build_mixed_sheets() -> bytes:
    """Non-tabular 'README' sheet + tabular 'Data' sheet."""

    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    readme["A1"] = "이 워크북 설명"
    readme["A2"] = "Data 시트에 실제 표가 있습니다."
    readme["A4"] = "문의: shkim@example.com"

    data = wb.create_sheet("Data")
    _write_rows(
        data,
        1,
        [
            ["item", "count", "unit"],
            ["apple", 3, "kg"],
            ["banana", 5, "kg"],
            ["cherry", 7, "box"],
            ["date", 9, "box"],
        ],
    )
    return _save_bytes(wb)


def build_cover_offset() -> bytes:
    """Cover sheet whose sparse text starts in column B (issue #3).

    Reproduces the original bug: text offset to column B leaves max_col=2, so
    the legacy ``max_col > 1`` gate misclassified it as tabular. Content-wise it
    is a single populated column -> expected non-tabular.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "표지"
    ws["B2"] = "2026년 영업 보고서"
    ws["B4"] = "작성: 영업팀"
    ws["B6"] = "기밀 — 외부 배포 금지"
    return _save_bytes(wb)


def build_cover_sparse() -> bytes:
    """Multi-column but scattered cover sheet (issue #3, density rule).

    Three cells in three different columns and rows -> populated_cols=3,
    populated_rows=3, filled=3, density=3/9=0.333. Above MIN_TABULAR_POPULATED_
    COLS but below NON_TABULAR_DENSITY_THRESHOLD -> expected non-tabular.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "표지"
    ws["B2"] = "분기 실적 요약"
    ws["E4"] = "2026-03-31"
    ws["C6"] = "재무팀"
    return _save_bytes(wb)


def build_sparse_real_table() -> bytes:
    """Genuine 4-column table with many missing cells (issue #3, density rule).

    Header + 5 data rows over columns A-D, with scattered blanks so
    populated_cols=4, populated_rows=6, filled=14, density=14/24=0.583. Above
    NON_TABULAR_DENSITY_THRESHOLD (0.5) -> expected tabular. Pins the density
    margin: raising the threshold past 0.583 would wrongly skip this real table.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(
        ws,
        1,
        [
            ["품목", "1월", "2월", "3월"],
            ["연필", 100, None, None],
            ["지우개", None, 50, None],
            ["공책", None, None, 210],
            ["펜", None, None, 90],
            ["자", 30, None, None],
        ],
    )
    return _save_bytes(wb)


def build_hidden_sheet() -> bytes:
    """Visible + hidden + veryHidden sheets (see ``FIXTURES``).

    Sheet order: 'Visible' (visible), 'Hidden' (sheet_state='hidden'),
    'VeryHidden' (sheet_state='veryHidden'). The latter two must report
    ``is_visible=False`` from the sheet enumerator.
    """

    wb = Workbook()
    visible = wb.active
    visible.title = "Visible"
    _write_rows(
        visible,
        1,
        [
            ["item", "qty", "unit"],
            ["apple", 3, "kg"],
            ["banana", 5, "kg"],
        ],
    )

    hidden = wb.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    _write_rows(hidden, 1, [["k", "v"], ["a", 1], ["b", 2]])

    very = wb.create_sheet("VeryHidden")
    very.sheet_state = "veryHidden"
    _write_rows(very, 1, [["k", "v"], ["c", 3], ["d", 4]])

    return _save_bytes(wb)


def build_blank_run_terminates() -> bytes:
    """Data block, a 2-row blank run terminator, then a noise block.

    Header row 1, data rows 2-5, blank rows 6-7 (BLANK_RUN terminator), then an
    unrelated noise label (row 9) and a trailing noise row (row 10). See
    ``FIXTURES`` for the exact 1-based boundary expectations used by Phase 3.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(
        ws,
        1,
        [
            ["name", "qty", "price"],  # row 1 header
            ["A-1", 10, 1.5],  # row 2
            ["A-2", 20, 2.5],  # row 3
            ["A-3", 30, 3.5],  # row 4
            ["A-4", 40, 4.5],  # row 5
            # rows 6-7 left empty -> BLANK_RUN terminator
        ],
    )
    # Noise block beyond the blank run (not part of the v1 table).
    ws["A9"] = "기타 메모"
    ws["A10"] = "Z-9"
    ws["B10"] = 99
    ws["C10"] = 9.9
    return _save_bytes(wb)


def build_interior_blank() -> bytes:
    """Data block with a single interior blank row (MEDIUM #4).

    Header row 1, data rows 2-3, a lone blank row 4 (below the BLANK_RUN of 2
    terminator threshold), then data rows 5-6. The blank row must be captured in
    ``skip_rows`` so it never leaks into the loaded frame as an all-NaN row. See
    ``FIXTURES`` for the exact 1-based boundary expectations.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(
        ws,
        1,
        [
            ["name", "qty", "price"],  # row 1 header
            ["A-1", 10, 1.5],  # row 2
            ["A-2", 20, 2.5],  # row 3
            [None, None, None],  # row 4 single interior blank
            ["A-3", 30, 3.5],  # row 5
            ["A-4", 40, 4.5],  # row 6
        ],
    )
    return _save_bytes(wb)


def build_empty_sheet() -> bytes:
    """A workbook with a single, completely empty sheet."""

    wb = Workbook()
    wb.active.title = "Sheet1"
    return _save_bytes(wb)


def build_header_only() -> bytes:
    """Header row 1 with no data rows beneath it."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(ws, 1, [["a", "b", "c"]])
    return _save_bytes(wb)


def build_no_header() -> bytes:
    """Pure homogeneous data with no header row (negative for header heuristic)."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(
        ws,
        1,
        [
            [1, 100, 1.1],
            [2, 200, 2.2],
            [3, 300, 3.3],
            [4, 400, 4.4],
            [5, 500, 5.5],
        ],
    )
    return _save_bytes(wb)


def build_large_table() -> bytes:
    """A large (5000-data-row) table for Phase 8 streaming/sampling proofs.

    Header row 1, :data:`LARGE_TABLE_DATA_ROWS` homogeneous data rows, then a
    single trailing ``합계`` grand-total row. The body is fully deterministic
    (no RNG) so the type sample and boundary detection are reproducible. The
    column types are chosen so the Type Profiler commits a definite type per
    column (``number`` / ``text`` / ``number``-as-flag), exercising the
    sampling path without ambiguity.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "id"
    ws["B1"] = "name"
    ws["C1"] = "amount"
    ws["D1"] = "flag"
    for i in range(LARGE_TABLE_DATA_ROWS):
        r = 2 + i  # 1-based data rows 2 .. 5001
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=f"item-{i:05d}")
        ws.cell(row=r, column=3, value=float(i) * 1.5)
        ws.cell(row=r, column=4, value=i % 2)
    # Trailing grand-total row (1-based row 2 + LARGE_TABLE_DATA_ROWS = 5002).
    total_row = 2 + LARGE_TABLE_DATA_ROWS
    ws.cell(row=total_row, column=1, value="합계")
    ws.cell(row=total_row, column=3, value=float(LARGE_TABLE_DATA_ROWS) * 1.5)
    return _save_bytes(wb)


def build_perf_100k() -> bytes:
    """100k-row performance-smoke workbook (plan v2 Phase 13 Step 4).

    **Not part of the default corpus** — excluded from ``FIXTURES`` /
    ``BUILDERS`` so corpus-parametrized tests never pay the build/inspect
    cost; only the ``@pytest.mark.slow`` memory smoke materializes it.

    Layout (1-based [D1]): header row 1 (A1:C1 = 'id','name','amount'),
    :data:`PERF_TABLE_DATA_ROWS` homogeneous data rows 2 ..
    ``PERF_TABLE_DATA_ROWS + 1``, columns A-C, no trailing total. Fully
    deterministic (no RNG) and timestamp-pinned via :func:`_save_bytes` like
    every corpus builder. Built with a write-only workbook so *generation*
    stays cheap (~1s) — a regular workbook's in-memory cell tree would dwarf
    the inspection budget being measured.

    Column-count calibration (measured, openpyxl 3.1.5 / Python 3.14.5): the
    structure-mode workbook (``read_only=False``, spec [D3] — merge/dimension
    metadata) materializes every cell at ~390-540 B each, so ``inspect()``'s
    tracemalloc peak is dominated by raw cell count, not by the sampled row
    data. At 100k x 4 columns the peak measured **216.8 MB — already over
    the spec §8 200 MB budget**; at 100k x 3 columns it is 171.4 MB. The
    fixture pins 3 columns so the §8 assertion is honestly satisfiable while
    still leaving ~30 MB of headroom to catch an accidental full
    materialization of row data in the analyzers (which would add far more).
    The 4-column overage is an open architecture-level finding (the [D3]
    full structure load breaches §8 at roughly >= 370k cells), recorded in
    the Phase 13 handoff — do not silently widen this fixture.
    """

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append(["id", "name", "amount"])
    for i in range(PERF_TABLE_DATA_ROWS):
        ws.append([i, f"item-{i:06d}", float(i) * 1.5])
    return _save_bytes(wb)


def build_multi_table_stacked() -> bytes:
    """Two vertically stacked tables split by a 2-row blank run (Phase 10).

    Table 1: header row 1 ('부서','인원','예산'), data rows 2-4, columns A-C.
    Rows 5-6 are fully blank (the BLANK_RUN separator). Table 2: header row 7
    ('제품명','단가','재고','비고'), data rows 8-10, columns A-D. See
    ``FIXTURES`` for the exact 1-based band/boundary expectations.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(
        ws,
        1,
        [
            ["부서", "인원", "예산"],  # row 1 header (table 1)
            ["영업", 12, 3400],  # row 2
            ["개발", 20, 5200],  # row 3
            ["관리", 5, 1100],  # row 4
            # rows 5-6 left fully empty -> BLANK_RUN band separator
        ],
    )
    _write_rows(
        ws,
        7,
        [
            ["제품명", "단가", "재고", "비고"],  # row 7 header (table 2)
            ["키보드", 30000, 15, "신품"],  # row 8
            ["마우스", 15000, 30, "신품"],  # row 9
            ["모니터", 210000, 7, "리퍼"],  # row 10
        ],
    )
    return _save_bytes(wb)


def build_stacked_uneven_width() -> bytes:
    """Stacked tables of uneven width: 3 columns over 8 columns (guard #1).

    Table 1 (narrow): header row 1 ('코드','명칭','수량'), data rows 2-4,
    columns A-C. Rows 5-6 blank (BLANK_RUN separator). Table 2 (wide): header
    row 7 (8 string labels), data rows 8-10, columns A-H. Exercises the
    score-denominator dilution guard (plan v2 §4 Task 10.2 guard 1): with a
    sheet-global ``max_col`` (8) denominator the narrow table's scores are
    diluted, so per-band analysis must use band-local column counts.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(
        ws,
        1,
        [
            ["코드", "명칭", "수량"],  # row 1 header (narrow table)
            ["K-01", "키트", 4],  # row 2
            ["K-02", "램프", 9],  # row 3
            ["K-03", "케이스", 2],  # row 4
            # rows 5-6 left fully empty -> BLANK_RUN band separator
        ],
    )
    _write_rows(
        ws,
        7,
        [
            # row 7 header (wide table, 8 columns)
            ["일자", "지점", "담당", "품목", "단가", "수량", "금액", "비고"],
            [_dt.date(2026, 3, 2), "서울", "김지원", "키트", 1000, 4, 4000, "정상"],  # row 8
            [_dt.date(2026, 3, 3), "부산", "이수민", "램프", 2000, 9, 18000, "정상"],  # row 9
            [_dt.date(2026, 3, 4), "대구", "박현우", "케이스", 1500, 2, 3000, "지연"],  # row 10
        ],
    )
    return _save_bytes(wb)


def build_title_between_tables() -> bytes:
    """Title band + two stacked tables + footnote band (W-A review HIGH).

    Row 1 title, rows 2-3 blank, table 1 rows 4-7 (header 4, data 5-7),
    rows 8-9 blank, table 2 rows 10-13 (header 10, data 11-13), rows 14-15
    blank, row 16 string footnote. See ``FIXTURES`` for the exact 1-based
    band/boundary expectations.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "2026년 1분기 부서별 집계"  # row 1: 1-row title band
    _write_rows(
        ws,
        4,
        [
            ["부서", "인원", "예산"],  # row 4 header (table 1)
            ["영업", 12, 3400],  # row 5
            ["개발", 20, 5200],  # row 6
            ["관리", 5, 1100],  # row 7
            # rows 8-9 left fully empty -> BLANK_RUN band separator
        ],
    )
    _write_rows(
        ws,
        10,
        [
            ["품목", "수량", "금액"],  # row 10 header (table 2)
            ["프린터", 3, 450000],  # row 11
            ["스캐너", 2, 380000],  # row 12
            ["복합기", 1, 520000],  # row 13
            # rows 14-15 left fully empty -> BLANK_RUN band separator
        ],
    )
    ws["A16"] = "주: 단위는 천원"  # row 16: 1-row string footnote band
    return _save_bytes(wb)


#: Cached formula results injected into the ``formulas_cached`` sheet XML,
#: keyed by A1 cell reference (the values Excel would cache after a recalc
#: save: ``=B{r}*C{r}`` over 수량/단가). Exported so tests assert the loaded
#: values from the single source of truth.
FORMULA_CACHE_VALUES: dict[str, int] = {
    "D2": 200,
    "D3": 600,
    "D4": 1200,
    "D5": 2000,
}

#: Zip member holding the single sheet of the ``formulas*`` fixtures.
_FORMULAS_SHEET_MEMBER = "xl/worksheets/sheet1.xml"


def _formulas_workbook() -> Workbook:
    """Shared layout of the two ``formulas*`` fixtures (Phase 12).

    Header row 1 (``품목,수량,단가,금액``), data rows 2-5 (columns A-D); the
    D column holds ``=B{r}*C{r}`` formulas. openpyxl serializes each formula
    cell as ``<c r="D{r}"><f>B{r}*C{r}</f><v /></c>`` — an *empty* cached
    value — so in data mode (``data_only=True``) the column reads all
    ``None`` (measured, openpyxl 3.1.5).
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(ws, 1, [["품목", "수량", "단가", "금액"]])
    body: list[tuple[str, int, int]] = [
        ("연필", 2, 100),  # row 2 -> D2 = =B2*C2 (would evaluate to 200)
        ("지우개", 3, 200),  # row 3 -> 600
        ("공책", 4, 300),  # row 4 -> 1200
        ("볼펜", 5, 400),  # row 5 -> 2000
    ]
    for offset, (item, qty, price) in enumerate(body):
        r = 2 + offset
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=qty)
        ws.cell(row=r, column=3, value=price)
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
    return wb


def build_formulas() -> bytes:
    """D-column ``=B*C`` formulas with no cached values (Phase 12 golden)."""

    return _save_bytes(_formulas_workbook())


def _inject_formula_caches(
    sheet_xml: bytes, caches: dict[str, int]
) -> bytes:
    """Insert a cached ``<v>`` result into each formula cell (plan v2 §6 probe 2).

    openpyxl never writes formula caches, so the ``as_value`` branch of the
    Formula Detector can only be exercised against a hand-crafted file: this
    rewrites the worksheet XML the way Excel would after a recalc save,
    turning ``<c r="D2"><f>B2*C2</f><v /></c>`` into
    ``<c r="D2"><f>B2*C2</f><v>200</v></c>``. The replacement is a pure,
    deterministic byte transform (no wall-clock input), so the cached fixture
    stays byte-stable across builds.

    Args:
        sheet_xml: The raw ``xl/worksheets/sheet1.xml`` payload.
        caches: ``{cell_ref: cached_numeric_result}`` to inject.

    Returns:
        The rewritten sheet XML.

    Raises:
        RuntimeError: A cell reference did not match exactly one formula cell
            (layout drift guard — fail loudly instead of silently producing a
            cacheless fixture that turns the as_value test into dead code).
    """

    for ref, value in caches.items():
        pattern = re.compile(
            (
                r'(<c r="{ref}"[^>]*>(?:<f[^>]*>[^<]*</f>|<f[^>]*/>))'
                r"(?:<v\s*/>|<v></v>)?"
            )
            .format(ref=re.escape(ref))
            .encode("utf-8")
        )
        sheet_xml, count = pattern.subn(
            rb"\g<1><v>" + str(value).encode("ascii") + rb"</v>", sheet_xml
        )
        if count != 1:
            raise RuntimeError(
                f"formula cache injection for {ref} matched {count} cells "
                f"(expected exactly 1); the formulas fixture layout drifted"
            )
    return sheet_xml


def build_formulas_cached() -> bytes:
    """The ``formulas`` layout with real cached ``<v>`` results (probe 2).

    Starts from the (already timestamp-normalized) :func:`build_formulas`
    bytes and repacks the zip with the sheet XML rewritten by
    :func:`_inject_formula_caches`; every member keeps the pinned
    ``_ZIP_EPOCH`` timestamp, so the output is byte-deterministic.
    """

    data = build_formulas()
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data)) as src:
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
            for info in src.infolist():
                payload = src.read(info.filename)
                if info.filename == _FORMULAS_SHEET_MEMBER:
                    payload = _inject_formula_caches(
                        payload, FORMULA_CACHE_VALUES
                    )
                new_info = zipfile.ZipInfo(
                    filename=info.filename, date_time=_ZIP_EPOCH
                )
                new_info.compress_type = zipfile.ZIP_DEFLATED
                new_info.external_attr = info.external_attr
                new_info.internal_attr = info.internal_attr
                new_info.create_system = info.create_system
                dst.writestr(new_info, payload)
    return out.getvalue()


def build_corrupt() -> bytes:
    """Truncated-zip bytes: first half of a valid .xlsx (BadZipFile)."""

    good = build_header_simple()
    half = max(1, len(good) // 2)
    return good[:half]


def build_encrypted() -> bytes:
    """Genuine password-protected .xlsx via msoffcrypto-tool (OLE2 container).

    Raises:
        RuntimeError: If ``msoffcrypto-tool`` is unavailable; the corpus
            cannot be produced without it (see module docstring / deviations).
    """

    try:
        import msoffcrypto  # noqa: PLC0415 - optional test-only dependency
    except ImportError as exc:  # pragma: no cover - defensive
        raise RuntimeError(
            "msoffcrypto-tool is required to generate the encrypted fixture; "
            "install it (see requirements.txt test deps)."
        ) from exc

    plain = io.BytesIO(build_header_simple())
    office = msoffcrypto.OfficeFile(plain)
    out = io.BytesIO()
    office.encrypt(ENCRYPTED_PASSWORD, out)
    return out.getvalue()


#: Ordered mapping of fixture id -> builder callable returning .xlsx bytes.
BUILDERS: dict[str, Callable[[], bytes]] = {
    "header_simple": build_header_simple,
    "header_offset": build_header_offset,
    "offset_plus_subtotals": build_offset_plus_subtotals,
    "merged_header": build_merged_header,
    "multi_level_header": build_multi_level_header,
    "multi_level_numeric_text": build_multi_level_numeric_text,
    "titled_multi_level": build_titled_multi_level,
    "stacked_multi_level": build_stacked_multi_level,
    "types_mixed": build_types_mixed,
    "left_margin_cols": build_left_margin_cols,
    "left_margin_with_subtotal": build_left_margin_with_subtotal,
    "mixed_sheets": build_mixed_sheets,
    "cover_offset": build_cover_offset,
    "cover_sparse": build_cover_sparse,
    "sparse_real_table": build_sparse_real_table,
    "hidden_sheet": build_hidden_sheet,
    "blank_run_terminates": build_blank_run_terminates,
    "interior_blank": build_interior_blank,
    "empty_sheet": build_empty_sheet,
    "header_only": build_header_only,
    "no_header": build_no_header,
    "large_table": build_large_table,
    "multi_table_stacked": build_multi_table_stacked,
    "stacked_uneven_width": build_stacked_uneven_width,
    "title_between_tables": build_title_between_tables,
    "formulas": build_formulas,
    "formulas_cached": build_formulas_cached,
    "corrupt": build_corrupt,
    "encrypted": build_encrypted,
}


# ---------------------------------------------------------------------------
# Corpus generation entry points
# ---------------------------------------------------------------------------


def fixtures_dir() -> Path:
    """Return the directory holding this module (the fixtures directory)."""

    return Path(__file__).resolve().parent


def generate_all(out_dir: str | Path | None = None) -> dict[str, Path]:
    """Generate the full fixture corpus into ``out_dir``.

    The corpus is deterministic and idempotent: regenerating overwrites the
    existing files with byte-stable content (modulo zip metadata).

    Args:
        out_dir: Destination directory. Defaults to this module's directory
            (``tests/fixtures``).

    Returns:
        Mapping of fixture id -> written :class:`pathlib.Path`.
    """

    target = Path(out_dir) if out_dir is not None else fixtures_dir()
    target.mkdir(parents=True, exist_ok=True)

    written: dict[str, Path] = {}
    for fixture_id, builder in BUILDERS.items():
        spec = FIXTURES[fixture_id]
        path = target / spec.filename
        path.write_bytes(builder())
        written[fixture_id] = path
    return written


def main(argv: list[str] | None = None) -> int:
    """CLI entry: ``python -m tests.fixtures.generate [out_dir]``."""

    args = list(sys.argv[1:] if argv is None else argv)
    out_dir = args[0] if args else None
    written = generate_all(out_dir)
    for fixture_id, path in written.items():
        print(f"{fixture_id}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
