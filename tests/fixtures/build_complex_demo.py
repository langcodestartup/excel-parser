"""Build a deliberately complex, realistic workbook to exercise the inspector.

One workbook, many hard features stacked together:
  - 표지        : scattered title text only           -> non-tabular (skipped)
  - 지역별매출  : 2 title rows + left margin col + 2-row merged/multi-level
                  header + numeric-text codes + dates + a 소계 subtotal row
  - 분기실적    : two stacked tables separated by a title band (uneven widths)
  - 계산        : formulas (so as_value / as_formula notes fire)
  - 원본(hidden): a hidden raw-dump sheet
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("complex_demo.xlsx")


def build() -> Path:
    wb = Workbook()

    # --- Sheet 1: 표지 (cover / non-tabular) -------------------------------
    cover = wb.active
    cover.title = "표지"
    cover["B2"] = "2026년 상반기 매출 보고서"
    cover["B4"] = "작성: 영업기획팀"
    cover["B6"] = "기밀 — 사내 한정"

    # --- Sheet 2: 지역별매출 (offset + margin + multi-level merged header) --
    s = wb.create_sheet("지역별매출")
    s["A1"] = "지역별 제품 매출 현황"          # title band row 1
    s["A2"] = "단위: 원 / 기준일 2026-06-01"   # title band row 2
    # row 4-5: two-row header. Top row merges describe groups.
    s["A4"] = "구분"          # left-margin label column (spans 2 header rows)
    s.merge_cells("A4:A5")
    s["B4"] = "제품 정보"
    s.merge_cells("B4:C4")    # merged group header over 제품코드 / 출시일
    s["D4"] = "실적"
    s.merge_cells("D4:E4")    # merged group header over 수량 / 매출액
    s["F4"] = "비고"
    s.merge_cells("F4:F5")
    s["B5"], s["C5"] = "제품코드", "출시일"
    s["D5"], s["E5"] = "수량", "매출액"

    rows = [
        ("서울", "00123", datetime(2026, 1, 5), 50, 20000.0, None),
        ("서울", "00481", datetime(2026, 2, 17), 30, 13500.5, "재입고"),
        ("부산", "00770", datetime(2026, 3, 2), 12, 4800.0, None),
        ("부산", "01099", datetime(2026, 3, 28), 0, 0.0, "단종"),
    ]
    r = 6
    for region, code, dt, qty, amt, memo in rows:
        s.cell(r, 1, region)
        s.cell(r, 2, code)     # leading-zero -> numeric_text (must stay string)
        s.cell(r, 3, dt)
        s.cell(r, 4, qty)
        s.cell(r, 5, amt)
        if memo is not None:
            s.cell(r, 6, memo)
        r += 1
    # subtotal row (should be detected/excluded, surfaced via notes)
    s.cell(r, 1, "소계")
    s.cell(r, 4, sum(x[3] for x in rows))
    s.cell(r, 5, sum(x[4] for x in rows))

    # --- Sheet 3: 분기실적 (two stacked tables, uneven width, title band) ---
    q = wb.create_sheet("분기실적")
    q["A1"] = "1분기"
    q["A2"], q["B2"], q["C2"] = "월", "매출", "신규고객"
    quarter1 = [("1월", 1000, 12), ("2월", 1100, 9), ("3월", 1250, 15)]
    rr = 3
    for m, sales, cust in quarter1:
        q.cell(rr, 1, m); q.cell(rr, 2, sales); q.cell(rr, 3, cust)
        rr += 1
    rr += 1  # blank separator row
    q.cell(rr, 1, "2분기"); rr += 1          # title band for second table
    # second table is narrower (uneven width on purpose)
    q.cell(rr, 1, "월"); q.cell(rr, 2, "매출"); rr += 1
    for m, sales in [("4월", 1320), ("5월", 1410)]:
        q.cell(rr, 1, m); q.cell(rr, 2, sales); rr += 1

    # --- Sheet 4: 계산 (formulas) ------------------------------------------
    c = wb.create_sheet("계산")
    c["A1"], c["B1"], c["C1"] = "항목", "단가", "수량"
    c["D1"] = "금액"
    data = [("연필", 500, 10), ("공책", 1200, 5), ("지우개", 300, 8)]
    for i, (name, price, qty) in enumerate(data, start=2):
        c.cell(i, 1, name); c.cell(i, 2, price); c.cell(i, 3, qty)
        c.cell(i, 4).value = f"=B{i}*C{i}"          # formula column
    c.cell(len(data) + 2, 1, "합계")
    c.cell(len(data) + 2, 4).value = f"=SUM(D2:D{len(data) + 1})"

    # --- Sheet 5: 원본 (hidden) --------------------------------------------
    h = wb.create_sheet("원본")
    h["A1"], h["B1"] = "key", "val"
    for i, (k, v) in enumerate([("a", 1), ("b", 2), ("c", 3)], start=2):
        h.cell(i, 1, k); h.cell(i, 2, v)
    h.sheet_state = "hidden"

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    p = build()
    print(f"wrote {p} ({p.stat().st_size} bytes)")
