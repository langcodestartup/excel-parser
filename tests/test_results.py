"""Result layer (extract / TableResult / WorkbookResult) tests — Phase 9."""
from __future__ import annotations

import datetime as dt
import decimal
import json
from pathlib import Path

import pandas as pd

from excel_inspector.results import (
    SCHEMA_VERSION,
    TableResult,
    WorkbookResult,
    _dedupe_columns,
    _jsonify_scalar,
)


# ---------------------------------------------------------------------------
# Scalar serialization rules (plan v2 §3.0, fixed contract)
# ---------------------------------------------------------------------------


def test_jsonify_scalar_rules() -> None:
    assert _jsonify_scalar(None) is None
    assert _jsonify_scalar(float("nan")) is None
    assert _jsonify_scalar(pd.NA) is None
    assert _jsonify_scalar(pd.NaT) is None
    assert _jsonify_scalar(pd.Timestamp("2026-01-05")) == "2026-01-05T00:00:00"
    assert _jsonify_scalar(dt.date(2026, 1, 5)) == "2026-01-05"
    assert _jsonify_scalar("00123") == "00123"  # numeric_text 문자열 유지
    import numpy as np

    assert _jsonify_scalar(np.int64(7)) == 7  # numpy -> python int
    assert type(_jsonify_scalar(np.int64(7))) is int


def test_jsonify_scalar_fallbacks_are_json_safe() -> None:
    """bytes/timedelta/Decimal -> explicit str fallback (review checklist).

    The contract forbids a silent ``json.dumps`` TypeError: every fallback
    output must round-trip through ``json.dumps`` unchanged.
    """

    import numpy as np

    assert _jsonify_scalar(b"raw-bytes") == "raw-bytes"
    assert _jsonify_scalar(dt.timedelta(hours=1, minutes=30)) == "1:30:00"
    assert _jsonify_scalar(pd.Timedelta(minutes=90)) == "0 days 01:30:00"
    assert _jsonify_scalar(np.timedelta64(90, "m")) == "1:30:00"
    assert _jsonify_scalar(decimal.Decimal("1.50")) == "1.50"
    for value in (
        b"raw-bytes",
        dt.timedelta(seconds=5),
        pd.Timedelta(seconds=5),
        decimal.Decimal("0.1"),
        np.timedelta64(5, "s"),
        np.datetime64("2026-01-05"),
    ):
        json.dumps(_jsonify_scalar(value))  # must not raise


# ---------------------------------------------------------------------------
# Column-name dedupe (JSON object keys must be unique)
# ---------------------------------------------------------------------------


def test_dedupe_columns() -> None:
    assert _dedupe_columns(["a", "b", "a", "a"]) == ["a", "b", "a.1", "a.2"]


def test_dedupe_columns_collision_with_existing_suffix() -> None:
    """["a", "a.1", "a"] must not mint a second "a.1" (review checklist).

    A naive counter would rename the trailing "a" to "a.1", colliding with the
    pre-existing "a.1" so the records dict silently drops a column.
    """

    out = _dedupe_columns(["a", "a.1", "a"])
    assert len(out) == len(set(out)), f"non-unique output: {out}"
    assert out[0] == "a" and out[1] == "a.1"


# ---------------------------------------------------------------------------
# TableResult / WorkbookResult serialization
# ---------------------------------------------------------------------------


def _toy_table() -> TableResult:
    df = pd.DataFrame({"지역": ["서울"], "수량": [50]})
    return TableResult(
        sheet_name="매출", table_id="매출!T1", dataframe=df,
        header_row=4, header_confidence=0.88, header_provenance="heuristic",
        columns=[], notes=[],
    )


def test_table_result_to_dict_shape() -> None:
    d = _toy_table().to_dict()
    assert d["table_id"] == "매출!T1"
    assert d["row_count"] == 1
    assert d["records"] == [{"지역": "서울", "수량": 50}]


def test_workbook_result_to_json_roundtrip() -> None:
    wr = WorkbookResult(file_path="/x.xlsx", sheets=[], warnings=[])
    parsed = json.loads(wr.to_json())
    assert parsed["schema_version"] == SCHEMA_VERSION
    assert parsed["file"] == "/x.xlsx"


def test_table_result_to_markdown_has_separator() -> None:
    md = _toy_table().to_markdown()
    lines = md.splitlines()
    assert lines[0].startswith("|") and "지역" in lines[0]
    assert set(lines[1].replace("|", "").replace(" ", "")) == {"-"}


def test_table_result_to_markdown_escapes_pipes_and_newlines() -> None:
    """Cell-borne '|' / newlines must not break the table grid (checklist)."""

    df = pd.DataFrame({"비고": ["a|b", "line1\nline2"], "값": [1, 2]})
    table = TableResult(
        sheet_name="S", table_id="S!T1", dataframe=df,
        header_row=1, header_confidence=1.0, header_provenance="manual",
        columns=[], notes=[],
    )
    md = table.to_markdown()
    lines = md.splitlines()
    # Grid intact: every line still has exactly 3 unescaped pipe separators.
    for line in lines:
        assert line.replace("\\|", "").count("|") == 3
    assert "a\\|b" in md
    assert "line1 line2" in md  # newline collapsed, row not broken


def test_table_result_to_dict_max_rows_truncates_records_only() -> None:
    """max_rows caps records but row_count stays the full table length."""

    df = pd.DataFrame({"v": [1, 2, 3, 4, 5]})
    table = TableResult(
        sheet_name="S", table_id="S!T1", dataframe=df,
        header_row=1, header_confidence=1.0, header_provenance="manual",
        columns=[], notes=[],
    )
    d = table.to_dict(max_rows=2)
    assert d["row_count"] == 5
    assert [r["v"] for r in d["records"]] == [1, 2]


def test_table_result_to_markdown_truncation_footer() -> None:
    df = pd.DataFrame({"v": list(range(25))})
    table = TableResult(
        sheet_name="S", table_id="S!T1", dataframe=df,
        header_row=1, header_confidence=1.0, header_provenance="manual",
        columns=[], notes=[],
    )
    md = table.to_markdown(max_rows=20)
    assert "… 5 more rows" in md


# ---------------------------------------------------------------------------
# extract() end-to-end over the real fixture corpus (Task 9.2)
# ---------------------------------------------------------------------------

from excel_inspector import extract


def test_extract_mixed_sheets(fixture_path) -> None:
    """mixed_sheets: 표 시트는 TableResult 1개, README는 skipped로 기록."""
    wr = extract(fixture_path("mixed_sheets"))
    assert [s.name for s in wr.sheets] == ["README", "Data"]
    readme, data = wr.sheets
    assert readme.skipped and readme.tables == [] and readme.skip_reason == "non-tabular"
    assert len(data.tables) == 1
    assert data.tables[0].table_id == "Data!T1"
    assert len(data.tables[0].dataframe) > 0


def test_extract_offset_plus_subtotals_records(fixture_path) -> None:
    """키스톤: 소계/합계가 records에 새지 않고 정확히 6행."""
    wr = extract(fixture_path("offset_plus_subtotals"))
    (table,) = wr.tables
    d = table.to_dict()
    assert d["row_count"] == 6
    assert sum(r["amount"] for r in d["records"]) == 590
    assert all("소계" not in str(r.values()) and "합계" not in str(r.values())
               for r in d["records"])


def test_extract_types_mixed_json_values(fixture_path) -> None:
    """numeric_text 선행 0 보존, date ISO 문자열, JSON 파싱 가능."""
    wr = extract(fixture_path("types_mixed"))
    parsed = json.loads(wr.to_json())
    recs = parsed["sheets"][0]["tables"][0]["records"]
    assert any(isinstance(r["code"], str) and r["code"].startswith("0") for r in recs)
    assert all(isinstance(r["date"], str) and r["date"][:4].isdigit() for r in recs)


def test_extract_headerless_override_col_names(fixture_path) -> None:
    """no_header + headerless override → 컬럼명 col_0..col_n."""
    from excel_inspector import InspectionOptions, SheetOverride
    opts = InspectionOptions(sheet_overrides={"Sheet1": SheetOverride(header_row=None)})
    wr = extract(fixture_path("no_header"), options=opts)
    (table,) = wr.tables
    assert list(table.dataframe.columns) == [f"col_{i}" for i in range(len(table.dataframe.columns))]


def test_extract_left_margin_with_subtotal_records(fixture_path) -> None:
    """L7 golden: the margin-shadowed '소계' row never leaks into records.

    The left-margin note on the subtotal row defeats a sheet-column-A keyword
    scan (plan v2 §3 review-checklist trap); with the data_left_col-anchored
    scan the loaded table is exactly the 5 data rows (qty sum 150) and no
    record carries the subtotal label or its price sum row.
    """
    wr = extract(fixture_path("left_margin_with_subtotal"))
    (table,) = wr.tables
    d = table.to_dict()
    assert d["row_count"] == 5
    assert sum(r["qty"] for r in d["records"]) == 150
    assert all("소계" not in str(r.values()) for r in d["records"])
    # The margin column A never enters the table (usecols C:E).
    assert list(table.dataframe.columns) == ["sku", "qty", "price"]


def test_extract_headerless_override_notes_dtype_skip(fixture_path) -> None:
    """L6 (plan v2 Phase 13 Step 2): the headerless note reaches the JSON.

    The aggregator records 'headerless sheet: dtype inference skipped' on the
    headerless plan; the result layer must surface it on the TableResult and
    in the serialized document, so the skipped dtype inference is visible end
    to end instead of silently yielding an empty dtype_map.
    """
    from excel_inspector import InspectionOptions, SheetOverride
    opts = InspectionOptions(sheet_overrides={"Sheet1": SheetOverride(header_row=None)})
    wr = extract(fixture_path("no_header"), options=opts)
    (table,) = wr.tables
    assert "headerless sheet: dtype inference skipped" in table.notes
    parsed = json.loads(wr.to_json())
    assert (
        "headerless sheet: dtype inference skipped"
        in parsed["sheets"][0]["tables"][0]["notes"]
    )


def test_extract_surfaces_excluded_subtotal_rows_in_notes(fixture_path) -> None:
    """Issue #2: excluded subtotal/total rows must surface in notes (No silent loss).

    offset_plus_subtotals drops subtotal rows 8 & 12 ('소계') and the grand-total
    row 13 ('합계') from the loaded frame (boundary detector ``skip_rows``). spec
    §8 forbids losing them silently, so each exclusion is recorded — with its
    1-based sheet row and label — on the table's ``notes`` and surfaces through
    the serialized JSON contract too.
    """
    wr = extract(fixture_path("offset_plus_subtotals"))
    (table,) = wr.tables
    assert "excluded subtotal/separator row at sheet row 8 (소계)" in table.notes
    assert "excluded subtotal/separator row at sheet row 12 (소계)" in table.notes
    assert "excluded subtotal/separator row at sheet row 13 (합계)" in table.notes
    parsed = json.loads(wr.to_json())
    notes = parsed["sheets"][0]["tables"][0]["notes"]
    assert "excluded subtotal/separator row at sheet row 8 (소계)" in notes


def test_extract_skip_rows_remove_clears_excluded_note(fixture_path) -> None:
    """[D2] × issue #2: keeping a subtotal via skip_rows_remove drops its note.

    ``skip_rows_remove=[8]`` re-includes subtotal row 8 in the loaded frame, so
    it is no longer excluded — its exclusion note must disappear while the still
    -dropped rows 12/13 keep theirs. (The note iterates the *post-override*
    interior skips, so the override is honored automatically.)
    """
    from excel_inspector import InspectionOptions, SheetOverride
    opts = InspectionOptions(
        sheet_overrides={"Sheet1": SheetOverride(header_row=4, skip_rows_remove=[8])}
    )
    wr = extract(fixture_path("offset_plus_subtotals"), options=opts)
    (table,) = wr.tables
    assert "excluded subtotal/separator row at sheet row 8 (소계)" not in table.notes
    assert "excluded subtotal/separator row at sheet row 12 (소계)" in table.notes
    assert "excluded subtotal/separator row at sheet row 13 (합계)" in table.notes
    # The kept subtotal row really is back in the loaded data.
    assert table.dataframe.astype(str).apply(
        lambda c: c.str.contains("소계")
    ).any().any()


def test_extract_surfaces_rows_above_header_in_notes(fixture_path) -> None:
    """Issue #8: non-empty rows absorbed above the detected header get a note.

    header_offset carries title rows 1-3 above the (correctly) detected header
    at row 4; Rule 1 absorbs them into ``skiprows`` so they never reach the
    frame. spec §8 forbids losing them silently — the dropped span surfaces on
    the table's ``notes`` and through the serialized JSON contract.
    """
    wr = extract(fixture_path("header_offset"))
    (table,) = wr.tables
    expected = (
        "rows above detected header not loaded: sheet rows 1-3 "
        "(header at row 4); use a header_row override if these are data rows"
    )
    assert expected in table.notes
    parsed = json.loads(wr.to_json())
    assert expected in parsed["sheets"][0]["tables"][0]["notes"]


def test_extract_small_mixed_table_detects_true_header(tmp_path) -> None:
    """Issue #8 acceptance: the true header wins in a small mixed-type table.

    Pre-fix, the §7.1 scoring picked the all-string data row 4 over the true
    header at row 1 (a 1-row lookahead window is trivially type-consistent),
    silently dropping rows 1-3. With the lookahead-evidence weighting the
    true header wins: all 4 data rows load and no rows-above-header note
    fires (nothing above the header was dropped).
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    rows = [
        ("항목", "값", "달성률"),
        ("총매출", 48300, "92%"),
        ("총수량", 318, "104%"),
        ("거래처수", "5개사", "100%"),
        ("반품건수", 1, "-"),
    ]
    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    path = tmp_path / "issue8_small_mixed.xlsx"
    wb.save(path)

    wr = extract(path)
    (table,) = wr.tables
    assert table.header_row == 1
    assert table.dataframe.shape == (4, 3)
    assert list(table.dataframe.columns) == ["항목", "값", "달성률"]
    assert not [
        n for n in table.notes if n.startswith("rows above detected header")
    ]


def test_extract_no_rows_above_note_across_bands(fixture_path) -> None:
    """A lower block's leading rows live in *other* bands -> no false note.

    title_between_tables stacks a title band, two table bands, and a footnote
    band. Each table's band starts at its own header row, so neither table
    drops rows above its header — the rows above table 2 belong to other
    bands (already surfaced via band-rejection warnings, plan v2 §4).
    """
    wr = extract(fixture_path("title_between_tables"))
    assert len(wr.tables) == 2
    for table in wr.tables:
        assert not [
            n for n in table.notes
            if n.startswith("rows above detected header")
        ]


def test_extract_json_is_deterministic(fixture_path) -> None:
    p = fixture_path("offset_plus_subtotals")
    assert extract(p).to_json() == extract(p).to_json()


def test_extract_non_string_header_labels_serialize(tmp_path) -> None:
    """Header cells [str, date, int] must not crash to_json() (P9 review, HIGH).

    A date header cell becomes a non-string pandas column label; the result
    layer must stringify it per the fixed serialization contract (ISO 8601),
    so json.loads(wr.to_json()) succeeds and every DataFrame column is str.
    """

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name", dt.date(2026, 1, 1), 2026])  # header: str / date / int
    ws.append(["alpha", 10, 1.5])
    ws.append(["beta", 20, 2.5])
    ws.append(["gamma", 30, 3.5])
    path = tmp_path / "nonstr_header.xlsx"
    wb.save(path)

    wr = extract(path)
    (table,) = wr.tables
    assert all(isinstance(c, str) for c in table.dataframe.columns)

    parsed = json.loads(wr.to_json())  # must not raise
    recs = parsed["sheets"][0]["tables"][0]["records"]
    assert recs, "no records extracted"
    keys = set(recs[0])
    assert "2026-01-01T00:00:00" in keys  # date label -> ISO 8601 string key
    assert "2026" in keys  # int label -> str key


def test_table_result_to_dict_with_non_string_columns_direct() -> None:
    """A directly-constructed TableResult never crashes on non-str labels.

    TableResult is a public dataclass: a caller can attach a DataFrame whose
    columns were never post-processed, so to_dict()/to_json() must stringify
    record keys themselves (P9 review, HIGH hardening).
    """

    df = pd.DataFrame([[1, 2]], columns=[dt.date(2026, 1, 1), 7])
    table = TableResult(
        sheet_name="S", table_id="S!T1", dataframe=df,
        header_row=1, header_confidence=1.0, header_provenance="manual",
        columns=[], notes=[],
    )
    parsed = json.loads(table.to_json())  # must not raise
    assert parsed["records"] == [{"2026-01-01": 1, "7": 2}]


def test_extract_relative_path_yields_absolute_file_field(
    fixture_path, monkeypatch
) -> None:
    """A relative input path is normalized to an absolute 'file' (P9 review, LOW)."""

    p = fixture_path("header_simple")
    monkeypatch.chdir(p.parent)
    wr = extract(p.name)  # relative path input
    assert Path(wr.file_path).is_absolute()
    assert Path(wr.file_path).name == p.name
    parsed = json.loads(wr.to_json())
    assert parsed["file"] == wr.file_path


# ---------------------------------------------------------------------------
# columns[].resolved_name — joinable with records keys (review MEDIUM #2)
# ---------------------------------------------------------------------------


def test_columns_expose_raw_and_resolved_names_merged_header(
    fixture_path,
) -> None:
    """merged_header: name stays raw; resolved_name equals the records key.

    The inspection-time names are ['이름', None, '점수'] (B1 is empty under
    the A1:B1 merge), but pandas labels the loaded empty header cell
    'Unnamed: 1' — pre-MEDIUM #2 the JSON columns could not be joined with
    the records. Both must now be exposed side by side.
    """

    wr = extract(fixture_path("merged_header"))
    (table,) = wr.tables
    d = table.to_dict()
    assert [c["name"] for c in d["columns"]] == ["이름", None, "점수"]
    assert [c["resolved_name"] for c in d["columns"]] == [
        "이름", "Unnamed: 1", "점수",
    ]
    assert set(d["records"][0]) == {c["resolved_name"] for c in d["columns"]}


def test_columns_resolved_name_flattened_multi_level(fixture_path) -> None:
    """multi_level_header: resolved_name carries the flattened '상위 / 하위'.

    The raw leaf names ('1월'..) stay on 'name' (inspection metadata); the
    flattened post-load names join the records keys positionally
    (ColumnProfile.index == 0-based selected-frame position [D5]).
    """

    wr = extract(fixture_path("multi_level_header"))
    (table,) = wr.tables
    d = table.to_dict()
    assert [c["name"] for c in d["columns"]] == ["1월", "2월", "3월", "4월"]
    assert [c["resolved_name"] for c in d["columns"]] == [
        "상반기 / 1월", "상반기 / 2월", "하반기 / 3월", "하반기 / 4월",
    ]
    assert set(d["records"][0]) == {c["resolved_name"] for c in d["columns"]}


def test_resolved_name_does_not_mutate_inspection_profiles(
    fixture_path,
) -> None:
    """TableResult copies the profiles — inspect() output stays untouched.

    build_workbook_result hands the *same* ColumnProfile instances from the
    inspection profile to TableResult; resolution must replace them with
    copies so the inspection domain never grows load-time state.
    """

    from excel_inspector import inspect

    profile = inspect(fixture_path("merged_header"))
    assert all(
        c.resolved_name is None for c in profile.sheets[0].columns
    )  # inspection-time default

    from excel_inspector.results import build_workbook_result

    wr = build_workbook_result(fixture_path("merged_header"), profile)
    (table,) = wr.tables
    assert [c.resolved_name for c in table.columns] == [
        "이름", "Unnamed: 1", "점수",
    ]
    # The inspection profile's own column objects were never mutated.
    assert all(c.resolved_name is None for c in profile.sheets[0].columns)


def test_resolved_names_join_records_keys_whole_corpus(
    openable_fixture: Path,
) -> None:
    """Every tabular fixture: records keys == the resolved_name set.

    The MEDIUM #2 acceptance assertion: whenever the profiled column count
    matches the loaded frame width, ``set(records[0].keys()) ==
    {c['resolved_name'] for c in columns}`` — i.e. the JSON columns metadata
    is positionally joinable with the records.
    """

    wr = extract(openable_fixture)
    parsed = json.loads(wr.to_json())
    for sheet in parsed["sheets"]:
        for tbl in sheet["tables"]:
            for col in tbl["columns"]:
                assert "name" in col and "resolved_name" in col
            if tbl["records"] and len(tbl["columns"]) == len(tbl["records"][0]):
                assert set(tbl["records"][0]) == {
                    c["resolved_name"] for c in tbl["columns"]
                }


def test_extract_every_openable_fixture_serializes(openable_fixture: Path) -> None:
    """extract() + to_json()/to_markdown() never crash over the whole corpus.

    Schema smoke: the JSON parses, declares schema v1.0, and every sheet entry
    carries the fixed key set (plan v2 §3.0 — shape is stable across phases).
    """

    wr = extract(openable_fixture)
    parsed = json.loads(wr.to_json())
    assert parsed["schema_version"] == SCHEMA_VERSION
    assert parsed["file"] == str(openable_fixture)
    assert isinstance(parsed["warnings"], list)
    for sheet in parsed["sheets"]:
        assert set(sheet) == {"name", "is_visible", "tables", "skipped", "skip_reason"}
        for tbl in sheet["tables"]:
            assert set(tbl) == {
                "table_id", "header_row", "header_confidence",
                "header_provenance", "columns", "row_count", "records", "notes",
            }
            assert tbl["row_count"] == len(tbl["records"])
    assert isinstance(wr.to_markdown(), str)


def test_offset_cover_sheet_skipped_not_empty_table(fixture_path) -> None:
    """issue #3: B열에서 시작하는 표지는 columns=[] 빈 테이블이 아니라 skip."""

    wr = extract(fixture_path("cover_offset"))
    assert [s.name for s in wr.sheets] == ["표지"]
    cover = wr.sheets[0]
    assert cover.skipped is True
    assert cover.skip_reason == "non-tabular"
    assert cover.tables == []


def test_sparse_cover_sheet_skipped(fixture_path) -> None:
    """issue #3: 다중 열이지만 희소한 표지도 skip(non-tabular)으로 끝난다."""

    wr = extract(fixture_path("cover_sparse"))
    cover = wr.sheets[0]
    assert cover.skipped is True and cover.skip_reason == "non-tabular"
    assert cover.tables == []
