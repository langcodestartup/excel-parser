"""Multi-level header loading tests — plan v2 Phase 11b (Task 11.2).

Three layers, mirroring the task steps:

* **Step 0 spike pins** — the pandas 3.0.3 behaviors the design depends on,
  measured (never assumed) against in-memory workbooks:

  (a) with an absolute ``header=[r0, r1]`` list, pandas forward-fills merged
      group cells, and a genuinely empty header cell becomes the
      ``Unnamed: N_level_M`` placeholder;
  (b) positional integer ``dtype`` keys stay **position**-valid under a list
      header [D5];
  (c) ``nrows`` counts the original rows consumed after the **last** header
      row (header rows are not in the budget; interior list-skips are);
  (c2) a list ``skiprows`` above the band breaks pandas' merged-cell header
      fill (fill happens at *pre-skip* absolute rows) — the measured pathology
      that forces the aggregator to emit **absolute** header indices and NOT
      absorb leading rows;
  (c3) ``usecols`` + multi-index header raises ``ValueError`` — the measured
      basis of the aggregator's usecols veto.

* **Aggregator unit tests** — Step 1: the contiguous merged band above a
  heuristic leaf header becomes ``ReadPlan.header = [band_top0, .., leaf0]``;
  non-contiguous bands / usecols conflicts / manual overrides conservatively
  keep the single leaf header.

* **Goldens** — Steps 2-4: ``multi_level_header.xlsx`` and
  ``multi_level_numeric_text.xlsx`` end-to-end through :func:`extract`,
  flattened ``"상위 / 하위"`` column names, record values, row counts, and the
  positional dtype_map key applied under ``header=list``. All expected values
  come from the ``FIXTURES`` descriptions (single source).

* **Stacked lower-band regression** (adversarial review MEDIUM #1) —
  ``stacked_multi_level.xlsx``: a lower band's group merges are classified
  ``body`` against the mirrored sheet header, so the aggregator's band-scoped
  path must re-classify them against the block's own header — otherwise the
  lower table's multi-level header is silently lost and a bogus forward-fill
  note is attached.
"""

from __future__ import annotations

import io
import json

import pandas as pd
import pytest
from openpyxl import Workbook

from excel_inspector import InspectionOptions, SheetOverride, extract, inspect
from excel_inspector.adapters.pandas_loader import read_plan_to_kwargs
from excel_inspector.aggregator import build_read_plan
from excel_inspector.models import MergeRegion, ReadPlan
from excel_inspector.results import (
    TableResult,
    _dedupe_columns,
    _flatten_column_tuple,
    _postprocess_dataframe,
    _stringify_label,
)

# ---------------------------------------------------------------------------
# Step 0 spike pins — measured pandas 3.0.3 behavior (in-memory workbooks)
# ---------------------------------------------------------------------------


def _two_level_workbook() -> io.BytesIO:
    """Group merges on row 1 (A1:B1 '상반기', C1:D1 '하반기'), leaves row 2."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "상반기"
    ws.merge_cells("A1:B1")
    ws["C1"] = "하반기"
    ws.merge_cells("C1:D1")
    rows = [
        ["1월", "2월", "3월", "4월"],
        [10, 20, 30, 40],
        [11, 21, 31, 41],
        [12, 22, 32, 42],
        [13, 23, 33, 43],
    ]
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _offset_band_workbook() -> io.BytesIO:
    """Title row 1, blank row 2, group row 3, leaf row 4, data rows 5-8.

    Row 7 is a '소계' subtotal so the interior-skip budget rule can be pinned
    under a list header.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "제목"
    ws["A3"] = "상반기"
    ws.merge_cells("A3:B3")
    ws["C3"] = "하반기"
    ws.merge_cells("C3:D3")
    rows = [
        ["1월", "2월", "3월", "4월"],  # row 4 leaf header
        [10, 20, 30, 40],  # row 5
        [11, 21, 31, 41],  # row 6
        ["소계", 41, 61, 81],  # row 7 subtotal
        [12, 22, 32, 42],  # row 8
    ]
    for r, row in enumerate(rows, start=4):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _numeric_text_group_workbook() -> io.BytesIO:
    """A1 left EMPTY (no group over the first column); group B1:C1 '그룹'."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B1"] = "그룹"
    ws.merge_cells("B1:C1")
    rows = [["코드", "x", "y"], ["007", 1, 2], ["012", 3, 4]]
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_spike_a_merged_group_labels_forward_filled() -> None:
    """(a) pandas fills merged group cells under an absolute header list."""

    df = pd.read_excel(
        _two_level_workbook(), sheet_name="Sheet1", engine="openpyxl",
        header=[0, 1],
    )
    assert list(df.columns) == [
        ("상반기", "1월"), ("상반기", "2월"), ("하반기", "3월"), ("하반기", "4월"),
    ]
    assert df.shape == (4, 4)


def test_spike_a_empty_group_cell_yields_unnamed_level_label() -> None:
    """(a) a genuinely empty group cell becomes 'Unnamed: N_level_M'."""

    df = pd.read_excel(
        _numeric_text_group_workbook(), sheet_name="Sheet1", engine="openpyxl",
        header=[0, 1],
    )
    assert list(df.columns) == [
        ("Unnamed: 0_level_0", "코드"), ("그룹", "x"), ("그룹", "y"),
    ]


def test_spike_b_positional_dtype_keys_valid_with_header_list() -> None:
    """(b) [D5] dtype={0: 'string'} types the first column under header=list."""

    df = pd.read_excel(
        _numeric_text_group_workbook(), sheet_name="Sheet1", engine="openpyxl",
        header=[0, 1], dtype={0: "string"},
    )
    assert str(df.dtypes.iloc[0]) == "string"
    assert list(df.iloc[:, 0]) == ["007", "012"]  # leading zeros survive


def test_spike_c_nrows_counts_rows_after_last_header_row() -> None:
    """(c) header rows never consume the nrows budget; data rows do."""

    for nrows, expected in ((2, [10, 11]), (4, [10, 11, 12, 13]),
                            (5, [10, 11, 12, 13])):
        df = pd.read_excel(
            _two_level_workbook(), sheet_name="Sheet1", engine="openpyxl",
            header=[0, 1], nrows=nrows,
        )
        assert list(df.iloc[:, 0]) == expected


def test_spike_c_interior_skiprows_consume_nrows_budget() -> None:
    """(c) interior list-skips below the band still consume the budget [D1].

    nrows=4 spans data rows 5-8 inclusive; the skipped subtotal row 7 is
    dropped from the output yet consumed from the budget, so the last data
    row (12) is still reached — identical to the v1 single-header rule.
    """

    df = pd.read_excel(
        _offset_band_workbook(), sheet_name="Sheet1", engine="openpyxl",
        header=[2, 3], skiprows=[6], nrows=4,
    )
    assert list(df.columns) == [
        ("상반기", "1월"), ("상반기", "2월"), ("하반기", "3월"), ("하반기", "4월"),
    ]
    assert list(df.iloc[:, 0]) == [10, 11, 12]


def test_spike_c2_list_skiprows_above_band_break_merge_fill() -> None:
    """(c2) the measured pathology behind the absolute-header decision.

    With ``skiprows=[0, 1]`` + post-skip ``header=[0, 1]``, pandas selects the
    right header rows but applies the merged-cell forward fill at *pre-skip*
    absolute rows 0-1, leaving 'Unnamed' gaps where the merge continued. The
    absolute form ``header=[2, 3]`` with no leading skips fills correctly —
    hence the aggregator never absorbs rows above a multi-level band.
    """

    broken = pd.read_excel(
        _offset_band_workbook(), sheet_name="Sheet1", engine="openpyxl",
        header=[0, 1], skiprows=[0, 1], nrows=2,
    )
    assert list(broken.columns) == [
        ("상반기", "1월"),
        ("Unnamed: 1_level_0", "2월"),
        ("하반기", "3월"),
        ("Unnamed: 3_level_0", "4월"),
    ]

    absolute = pd.read_excel(
        _offset_band_workbook(), sheet_name="Sheet1", engine="openpyxl",
        header=[2, 3], nrows=2,
    )
    assert list(absolute.columns) == [
        ("상반기", "1월"), ("상반기", "2월"), ("하반기", "3월"), ("하반기", "4월"),
    ]


def test_spike_c3_usecols_with_multi_index_header_raises() -> None:
    """(c3) pandas rejects usecols + multi-index header — the veto basis."""

    with pytest.raises(ValueError, match="usecols"):
        pd.read_excel(
            _two_level_workbook(), sheet_name="Sheet1", engine="openpyxl",
            header=[0, 1], usecols="A:D",
        )


# ---------------------------------------------------------------------------
# Step 1 — aggregator: contiguous merged band -> ReadPlan.header list
# ---------------------------------------------------------------------------


def test_plan_header_list_for_multi_level_fixture(fixture_path) -> None:
    """multi_level_header: header=[0, 1], leading rows NOT absorbed.

    FIXTURES golden: band rows 1-2 (0-based [0, 1]), skiprows=[], nrows=4,
    no usecols, empty dtype_map (all columns number).
    """

    sheet = inspect(fixture_path("multi_level_header")).sheets[0]
    plan = sheet.read_plan
    assert plan is not None
    assert plan.header == [0, 1]
    assert plan.skiprows == []
    assert plan.nrows == 4
    assert plan.usecols is None
    assert plan.dtype_map == {}


def test_mirror_block_plan_carries_header_list(fixture_path) -> None:
    """The mirror block's independently-computed plan derives the same list."""

    sheet = inspect(fixture_path("multi_level_header")).sheets[0]
    assert len(sheet.blocks) == 1
    block_plan = sheet.blocks[0].read_plan
    assert block_plan is not None
    assert block_plan.header == [0, 1]
    assert block_plan == sheet.read_plan
    assert block_plan is not sheet.read_plan  # independently computed


def test_offset_band_absolute_indices_not_absorbed(
    sheet_profile_factory,
) -> None:
    """A band away from row 1: absolute 0-based indices, no leading skips.

    Band rows 3-4 (1-based) -> header=[2, 3]; rows 1-2 above the band stay
    unskipped (pandas ignores rows above the first header row); the interior
    subtotal skip at row 7 converts to 0-based 6 as usual [D1].
    """

    profile = sheet_profile_factory(
        max_row=8,
        max_col=4,
        header_row=4,
        header_provenance="heuristic",
        is_multi_level_header=True,
        merges=[
            MergeRegion(range="A3:B3", kind="header"),
            MergeRegion(range="C3:D3", kind="header"),
        ],
        data_start_row=5,
        data_end_row=8,
        skip_rows=[7],
    )
    warnings: list[str] = []
    plan = build_read_plan(profile, None, warnings)
    assert plan.header == [2, 3]
    assert plan.skiprows == [6]
    assert plan.nrows == 4
    assert warnings == []


def test_vertical_merge_overlapping_header_counts_as_band(
    sheet_profile_factory,
) -> None:
    """A merge spanning rows 1-2 over a row-2 leaf header forms the band."""

    profile = sheet_profile_factory(
        max_row=4,
        max_col=3,
        header_row=2,
        header_provenance="heuristic",
        is_multi_level_header=True,
        merges=[MergeRegion(range="A1:A2", kind="header")],
        data_start_row=3,
        data_end_row=4,
    )
    plan = build_read_plan(profile, None, [])
    assert plan.header == [0, 1]
    assert plan.skiprows == []


def test_non_contiguous_band_keeps_single_header_with_warning(
    sheet_profile_factory,
) -> None:
    """A gap between the merged band and the leaf header vetoes the list."""

    profile = sheet_profile_factory(
        max_row=6,
        max_col=4,
        header_row=3,  # merged row 1, bare row 2 -> gap
        header_provenance="heuristic",
        is_multi_level_header=True,
        merges=[MergeRegion(range="A1:B1", kind="header")],
        data_start_row=4,
        data_end_row=6,
    )
    warnings: list[str] = []
    plan = build_read_plan(profile, None, warnings)
    assert plan.header == 0
    assert plan.skiprows == [0, 1]
    assert len(warnings) == 1
    assert "not contiguous" in warnings[0]


def test_disconnected_title_merge_keeps_contiguous_band_with_warning(
    sheet_profile_factory,
) -> None:
    """A merged title row above a gap no longer vetoes the band (issue #7).

    Merged title row 1, blank row 2, group merges row 3, leaf header row 4:
    the maximal contiguous run ending at the leaf (rows 3-4) still loads as
    the multi-level band ``header=[2, 3]``; only the disconnected title row
    is dropped from the band, with a visible warning (spec §8 no silent
    loss) — NOT the previous all-or-nothing fallback to the bare leaf row
    that broke the vertical-merge anchor names into ``Unnamed: N``.
    """

    profile = sheet_profile_factory(
        max_row=6,
        max_col=4,
        header_row=4,
        header_provenance="heuristic",
        is_multi_level_header=True,
        merges=[
            MergeRegion(range="A1:D1", kind="header"),  # title banner, row 1
            MergeRegion(range="A3:B3", kind="header"),
            MergeRegion(range="C3:D3", kind="header"),
        ],
        data_start_row=5,
        data_end_row=6,
    )
    warnings: list[str] = []
    plan = build_read_plan(profile, None, warnings)
    assert plan.header == [2, 3]
    assert plan.skiprows == []
    assert len(warnings) == 1
    assert "excluded from the multi-level header" in warnings[0]
    assert "[1]" in warnings[0]


def test_usecols_conflict_keeps_single_header_with_warning(
    sheet_profile_factory,
) -> None:
    """pandas rejects usecols + multi-index header -> conservative fallback."""

    profile = sheet_profile_factory(
        max_row=5,
        max_col=4,
        header_row=2,
        header_provenance="heuristic",
        is_multi_level_header=True,
        merges=[MergeRegion(range="C1:D1", kind="header")],
        data_start_row=3,
        data_end_row=5,
        data_left_col=3,
        data_right_col=4,
    )
    warnings: list[str] = []
    plan = build_read_plan(profile, None, warnings)
    assert plan.header == 0
    assert plan.usecols == "C:D"
    assert plan.skiprows == [0]
    assert len(warnings) == 1
    assert "usecols" in warnings[0]


def test_manual_header_override_keeps_single_header_silently(
    fixture_path,
) -> None:
    """[D2] a manual header_row is authoritative — never widened to a list."""

    options = InspectionOptions(
        sheet_overrides={"Sheet1": SheetOverride(header_row=2)}
    )
    profile = inspect(fixture_path("multi_level_header"), options)
    sheet = profile.sheets[0]
    assert sheet.header_provenance == "manual"
    assert sheet.is_multi_level_header is True
    plan = sheet.read_plan
    assert plan is not None
    assert plan.header == 0
    assert plan.skiprows == [0]
    assert not any("multi-level" in w for w in profile.open_errors)


def test_headerless_override_unaffected_by_multi_flag(
    sheet_profile_factory,
) -> None:
    """An explicit headerless declaration wins over the multi-level flag."""

    options = InspectionOptions(
        sheet_overrides={"Sheet1": SheetOverride(header_row=None)}
    )
    profile = sheet_profile_factory(
        max_row=4,
        max_col=4,
        header_row=2,
        header_provenance="heuristic",
        is_multi_level_header=True,
        merges=[MergeRegion(range="A1:B1", kind="header")],
    )
    plan = build_read_plan(profile, options, [])
    assert plan.header is None


def test_body_only_merges_never_trigger_header_list(
    sheet_profile_factory,
) -> None:
    """kind='body' merges above a block header add no band (silent single).

    This is the band-scoped lower-block situation: its merges were classified
    against the sheet header, so no 'header' merge exists in scope and the
    derivation finds no evidence — single header, no warning.
    """

    profile = sheet_profile_factory(
        max_row=6,
        max_col=3,
        header_row=2,
        header_provenance="heuristic",
        is_multi_level_header=True,
        merges=[MergeRegion(range="A1:B1", kind="body")],
        data_start_row=3,
        data_end_row=6,
    )
    warnings: list[str] = []
    plan = build_read_plan(profile, None, warnings)
    assert plan.header == 0
    assert plan.skiprows == [0]
    assert warnings == []


# ---------------------------------------------------------------------------
# Step 2 — adapter: the header list passes through verbatim
# ---------------------------------------------------------------------------


def test_read_plan_to_kwargs_passes_header_list_verbatim() -> None:
    """read_plan_to_kwargs needs no change for list headers (Task 11.2 §2)."""

    plan = ReadPlan(sheet_name="Sheet1", header=[0, 1], skiprows=[], nrows=4)
    kwargs = read_plan_to_kwargs(plan)
    assert kwargs["header"] == [0, 1]
    assert "usecols" not in kwargs  # None -> omitted (all columns)
    assert kwargs["nrows"] == 4


# ---------------------------------------------------------------------------
# Step 3 — results: MultiIndex flattening ("상위 / 하위")
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("parts", "expected"),
    [
        (("상반기", "1월"), "상반기 / 1월"),
        (("Unnamed: 0_level_0", "코드"), "코드"),  # unfilled group cell
        (("그룹", "Unnamed: 2_level_1"), "그룹"),  # unfilled leaf cell
        (("Unnamed: 3", "y"), "y"),  # flat-style placeholder
        ((None, "x"), "x"),  # missing group level
        (("", "리프"), "리프"),  # empty-string group level
        (("Unnamed: 0_level_0", "Unnamed: 0_level_1"), ""),  # nothing real
        (("A", "b", "c"), "A / b / c"),  # 3-level join
    ],
)
def test_flatten_column_tuple_rules(parts: tuple, expected: str) -> None:
    assert _flatten_column_tuple(parts) == expected


def test_postprocess_flattens_and_dedupes_multiindex() -> None:
    """Flattened duplicates get '.N' suffixes via _dedupe_columns (spike d)."""

    df = pd.DataFrame(
        [[1, 2, 3]],
        columns=pd.MultiIndex.from_tuples(
            [("G", "a"), ("G", "a"), ("Unnamed: 2_level_0", "b")]
        ),
    )
    plan = ReadPlan(sheet_name="Sheet1", header=[0, 1])
    out = _postprocess_dataframe(df, plan)
    assert list(out.columns) == ["G / a", "G / a.1", "b"]


def test_dedupe_columns_on_flattened_names() -> None:
    """(d) _dedupe_columns composes cleanly with flattened names."""

    flat = [_flatten_column_tuple(t) for t in
            [("G", "a"), ("G", "a"), ("", "b")]]
    assert _dedupe_columns(flat) == ["G / a", "G / a.1", "b"]


def test_stringify_label_flattens_tuple_for_direct_table_result() -> None:
    """A directly-built TableResult with a MultiIndex df gets clean JSON keys."""

    df = pd.DataFrame(
        [[10, 20]],
        columns=pd.MultiIndex.from_tuples([("상위", "하위"), ("상위", "둘째")]),
    )
    assert _stringify_label(("상위", "하위")) == "상위 / 하위"
    table = TableResult(
        sheet_name="S", table_id="S!T1", dataframe=df, header_row=2,
        header_confidence=1.0, header_provenance="heuristic", columns=[],
    )
    records = table.to_dict()["records"]
    assert records == [{"상위 / 하위": 10, "상위 / 둘째": 20}]


# ---------------------------------------------------------------------------
# Step 4 — goldens: extract() end-to-end (FIXTURES single source)
# ---------------------------------------------------------------------------


def test_extract_multi_level_header_golden(fixture_path) -> None:
    """multi_level_header: flattened columns, 4 rows, fixed column sums."""

    wr = extract(fixture_path("multi_level_header"))
    (table,) = wr.tables
    df = table.dataframe
    assert list(df.columns) == [
        "상반기 / 1월", "상반기 / 2월", "하반기 / 3월", "하반기 / 4월",
    ]
    assert table.header_row == 2  # 1-based leaf header (inspection domain)
    assert table.header_provenance == "heuristic"
    assert len(df) == 4
    assert [int(df[c].sum()) for c in df.columns] == [46, 86, 126, 166]

    d = table.to_dict()
    assert d["row_count"] == 4
    assert d["records"][0] == {
        "상반기 / 1월": 10, "상반기 / 2월": 20, "하반기 / 3월": 30, "하반기 / 4월": 40,
    }
    assert wr.warnings == []


def test_extract_multi_level_header_json_deterministic(fixture_path) -> None:
    """Two extract() runs serialize to identical JSON (flatten is stable)."""

    path = fixture_path("multi_level_header")
    first = extract(path).to_json()
    assert first == extract(path).to_json()
    keys = list(json.loads(first)["sheets"][0]["tables"][0]["records"][0])
    assert keys == ["상반기 / 1월", "상반기 / 2월", "하반기 / 3월", "하반기 / 4월"]


def test_extract_multi_level_numeric_text_dtype_positional(
    fixture_path,
) -> None:
    """[D5] the positional dtype_map key applies under header=list, end-to-end.

    FIXTURES golden: dtype_map={'0': 'string'} types the FIRST selected
    column ('기본 / 코드'), so the digit strings keep their leading zeros all
    the way into the records.
    """

    path = fixture_path("multi_level_numeric_text")
    sheet = inspect(path).sheets[0]
    plan = sheet.read_plan
    assert plan is not None
    assert plan.header == [0, 1]
    assert plan.skiprows == []
    assert plan.nrows == 3
    assert plan.dtype_map == {"0": "string"}

    wr = extract(path)
    (table,) = wr.tables
    df = table.dataframe
    assert list(df.columns) == [
        "기본 / 코드", "기본 / 수량", "실적 / 단가", "실적 / 금액",
    ]
    assert str(df.dtypes.iloc[0]) == "string"
    assert list(df["기본 / 코드"]) == ["007", "012", "034"]
    assert int(df["실적 / 금액"].sum()) == 4400

    records = json.loads(table.to_json())["records"]
    assert [r["기본 / 코드"] for r in records] == ["007", "012", "034"]


def test_extract_titled_multi_level_records_keys_match_columns(
    fixture_path,
) -> None:
    """[Issue #7 golden] a merged title banner does not break the band.

    End-to-end through extract(): the disconnected title row is excluded
    (with a visible warning) while rows 3-4 still load as the multi-level
    band, so records keys carry the flattened names — vertical-merge anchors
    included — and match ``columns[].resolved_name`` positionally. The
    numeric_text column under the band keeps its leading zeros [D5].
    """

    wr = extract(fixture_path("titled_multi_level"))
    (table,) = wr.tables
    assert table.header_row == 4  # 1-based leaf header (inspection domain)

    expected = [
        "지역", "제품코드",
        "1분기 / 1월", "1분기 / 2월", "1분기 / 3월",
        "2분기 / 4월", "2분기 / 5월", "2분기 / 6월",
    ]
    d = table.to_dict()
    assert [c["resolved_name"] for c in d["columns"]] == expected
    assert list(d["records"][0].keys()) == expected
    assert d["records"][0]["지역"] == "서울"
    assert [r["제품코드"] for r in d["records"]] == ["00123", "00789"]
    assert any(
        "excluded from the multi-level header" in w for w in wr.warnings
    )


# ---------------------------------------------------------------------------
# Stacked lower-band regression — adversarial review MEDIUM #1
# (band-scoped merges re-classified against the block's own header)
# ---------------------------------------------------------------------------


def test_stacked_lower_block_reclassifies_merges_to_header_band(
    fixture_path,
) -> None:
    """T2's group merges promote its multi-level header despite 'body' labels.

    The Merge Analyzer classified the row-7 group merges against the mirrored
    sheet header (row 1, blocks[0]) as kind='body' — the counterexample's
    precondition, pinned below. The aggregator's band-scoped path must
    re-classify them against the block's OWN header (row 8), yielding the
    FIXTURES golden plan: header=[6, 7] (0-based absolute band rows 7-8),
    skiprows=[], nrows=3, no usecols, empty dtype_map — and (a) NO bogus
    body-merge forward-fill note on either block's plan.
    """

    sheet = inspect(fixture_path("stacked_multi_level")).sheets[0]
    # Precondition (the misclassification under test): sheet-level merges are
    # anchored on the mirrored row-1 header, so both group merges read 'body'.
    assert {(m.range, m.kind) for m in sheet.merges} == {
        ("A7:B7", "body"),
        ("C7:D7", "body"),
    }

    assert len(sheet.blocks) == 2
    top, bottom = sheet.blocks
    assert top.read_plan is not None
    assert top.header_row == 1
    assert top.read_plan.header == 0
    assert top.read_plan.nrows == 3

    assert bottom.read_plan is not None
    assert bottom.header_row == 8  # 1-based leaf header (inspection domain)
    assert bottom.read_plan.header == [6, 7]
    assert bottom.read_plan.skiprows == []
    assert bottom.read_plan.nrows == 3
    assert bottom.read_plan.usecols is None
    assert bottom.read_plan.dtype_map == {}

    # (a) The misclassified group merges must not leak forward-fill notes.
    for plan in (top.read_plan, bottom.read_plan):
        assert not any(note.startswith("body merge") for note in plan.notes)


def test_extract_stacked_multi_level_golden(fixture_path) -> None:
    """(b) the lower table loads with flattened multi-level columns.

    FIXTURES golden: T1 stays the flat 4-column table; T2 flattens to
    '상반기 / 1월' .. '하반기 / 4월' with per-column sums 33/63/93/123 — the
    multi-level header promotion is no longer silently lost (MEDIUM #1).
    """

    wr = extract(fixture_path("stacked_multi_level"))
    t1, t2 = wr.tables
    assert t1.table_id == "Sheet1!T1"
    assert list(t1.dataframe.columns) == ["부서", "인원", "예산", "비고"]
    assert len(t1.dataframe) == 3

    assert t2.table_id == "Sheet1!T2"
    assert t2.header_row == 8
    df = t2.dataframe
    assert list(df.columns) == [
        "상반기 / 1월", "상반기 / 2월", "하반기 / 3월", "하반기 / 4월",
    ]
    assert len(df) == 3
    assert [int(df[c].sum()) for c in df.columns] == [33, 63, 93, 123]

    # No wrong body-merge note reaches either TableResult.
    for table in (t1, t2):
        assert not any(note.startswith("body merge") for note in table.notes)
